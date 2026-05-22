import os 
from pathlib import Path
import time
import numpy as np
import sys
cur_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(cur_dir)
folder_utils = os.path.join(project_root,'utils')
sys.path.insert(0,folder_utils)
from visualize import draw, draw_step_response_analysis
from topp import TOPP
import csv
from collections import deque
import threading
import matplotlib.pyplot as plt
import yaml
from uuid import uuid4
MAX_POINTS = 1000
CONTROL_HZ = 20
TOL = 1e-3
EXTRA_TIME = 0.2 # s
# decorator
def hide_ui_while(func):
    """
    装饰器：执行函数时隐藏UI，执行完再恢复
    假设函数所在对象 self 有 self.g.ui 属性
    """
    def wrapper(self, *args, **kwargs):
        ui = getattr(self.g, "ui", None)
        hidden_by_wrapper = False
        if ui is not None and getattr(ui, "show_ui", False):
            ui.simulate_key("h")
            hidden_by_wrapper = True
        try:
            result = func(self, *args, **kwargs)
        finally:
            if hidden_by_wrapper:
                ui.simulate_key("\n")
        return result
    return wrapper

class MotionModule:
    def __init__(self,guide):
        self.g = guide
        self.manual_control_active = False
        self.manual_control_step = 0.00005
        self.max_acc = 0.1
        self.max_vel = 0.1
        self.current_output_dir = None

    def _get_output_dir(self):
        out_dir = self.current_output_dir or os.path.join(project_root, "logs")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def _move_log_file_to_output_dir(self, filename: str):
        if not filename:
            return filename
        out_dir = self._get_output_dir()
        src = filename if os.path.isabs(filename) else os.path.join(project_root, "logs", filename)
        dst = os.path.join(out_dir, os.path.basename(filename))
        if os.path.abspath(src) != os.path.abspath(dst) and os.path.exists(src):
            os.replace(src, dst)
        return os.path.basename(dst)

    def _coerce_scalar(self, value):
        if value is None:
            return None
        if isinstance(value, (list, tuple)):
            if not value:
                return None
            value = value[0]
        if isinstance(value, np.ndarray):
            if value.size == 0:
                return None
            value = value.reshape(-1)[0]
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _move_to_target(self, part: str, pos_name: str, target: float, timeout_s: float = 3.0):
        target_arr = np.array([target], dtype=float)
        start_time = time.monotonic()
        while True:
            with self.g.feedback_lock:
                cur_q = getattr(self.g.feedbackData, pos_name)
            if np.max(np.abs(np.array(cur_q, dtype=float) - target_arr)) <= TOL:
                return True
            if time.monotonic() - start_time > timeout_s:
                return False
            self.g.robot.set_actions({part: {"type": "position", "position": target_arr.tolist()}})
            time.sleep(1 / CONTROL_HZ)

    def _get_feedback_scalar(self, name: str):
        with self.g.feedback_lock:
            raw = getattr(self.g.feedbackData, name, None)
        return self._coerce_scalar(raw)

    def _get_loadcell_channels(self):
        with self.g.feedback_lock:
            force_tip = list(getattr(self.g.feedbackData, "force_tip", [None, None]))
        ch0 = self._coerce_scalar(force_tip[0] if len(force_tip) > 0 else None)
        ch1 = self._coerce_scalar(force_tip[1] if len(force_tip) > 1 else None)
        return ch0, ch1

    def _format_loadcell_channels(self):
        ch0, ch1 = self._get_loadcell_channels()
        ch0_text = "None" if ch0 is None else f"{ch0:.4f}"
        ch1_text = "None" if ch1 is None else f"{ch1:.4f}"
        return f"通道0={ch0_text} N, 通道1={ch1_text} N"

    def _show_loadcell_channels_for(self, seconds: float, prefix: str = "当前一维力"):
        end_time = time.monotonic() + seconds
        while time.monotonic() < end_time:
            print(f"\r{prefix}: {self._format_loadcell_channels()}    ", end="", flush=True)
            time.sleep(0.1)
        print()


    def safe_get_feedback_snapshot(self, pos_name: str = "gripper_pos"):
        with self.g.feedback_lock:
            fb = self.g.feedbackData
            snapshot = {
                "gripper_pos": self._coerce_scalar(getattr(fb, pos_name, None)),
                "gripper_vel": self._coerce_scalar(getattr(fb, "gripper_vel", None)),
                "gripper_torque": self._coerce_scalar(getattr(fb, "gripper_torque", None)),
                "real_distance": self._coerce_scalar(getattr(fb, "real_distance", None)),
                "rb_time": self._coerce_scalar(getattr(fb, "rb_time", None)),
                "laser_status": getattr(fb, "laser_status", None),
            }
        return snapshot

    def _calibration_csv_headers(self):
        return [
            "timestamp",
            "run_id",
            "gripper_name",
            "direction",
            "step_index",
            "command_pos_rad",
            "gripper_pos",
            "gripper_opening_distance",
            "gripper opening distance",
            "distance",
            "delta_pos_from_prev",
            "gripper_vel",
            "gripper_torque",
            "rb_time",
            "laser_status",
            "limit_reason",
        ]

    def append_csv_row(self, csv_path: str, row_dict: dict):
        headers = self._calibration_csv_headers()
        csv_dir = os.path.dirname(csv_path)
        if csv_dir:
            os.makedirs(csv_dir, exist_ok=True)
        file_exists = os.path.exists(csv_path)
        write_header = (not file_exists) or os.path.getsize(csv_path) == 0

        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            if write_header:
                writer.writeheader()
            writer.writerow({k: row_dict.get(k) for k in headers})

    def detect_mechanical_limit_by_stall(
        self,
        delta_history: deque,
        pos_history: deque,
        threshold: float,
        consecutive_count: int,
        window_motion_threshold: float,
    ) -> bool:
        if consecutive_count <= 0 or len(delta_history) < consecutive_count:
            return False

        tail_delta = list(delta_history)[-consecutive_count:]
        if not all((v is not None) and (v <= threshold) for v in tail_delta):
            return False

        if len(pos_history) < (consecutive_count + 1):
            return False
        tail_pos = list(pos_history)[-(consecutive_count + 1):]
        motion_span = max(tail_pos) - min(tail_pos)
        return motion_span <= window_motion_threshold

    def _format_timestamp_ms(self):
        now = time.time()
        return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(now)) + f".{int((now % 1) * 1000):03d}"

    def collect_gripper_calibration_data(
        self,
        part: str,
        pos_name: str = "gripper_pos",
        direction: str = "open",
        step_rad: float = 0.002,
        settle_delay_s: float = 1.0,
        control_interval_s: float = 0.0,
        stall_delta_threshold: float = 5e-5,
        stall_consecutive_required: int = 20,
        stall_window_motion_threshold: float = 2e-4,
        min_steps_before_stall: int = 60,
        max_steps: int = 20000,
        max_duration_s: float = 0.0,
        csv_path: str = None,
    ):
        direction = str(direction).strip().lower()
        if direction not in ("open", "close"):
            raise ValueError("direction must be 'open' or 'close'")
        if step_rad <= 0:
            raise ValueError("step_rad must be > 0")

        if csv_path is None:
            csv_path = os.path.join(project_root, "logs", "gripper_calibration_dataset.csv")

        try:
            self.g.robot.send_command(part, {"command": "set_control_mode", "mode": "position"})
        except Exception:
            pass

        direction_sign = 1.0 if direction == "open" else -1.0
        run_id = f"{time.strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}"
        start_t = time.monotonic()
        step_index = 0
        prev_actual_pos = None
        delta_history = deque(maxlen=max(1, int(stall_consecutive_required)))
        pos_history = deque(maxlen=max(2, int(stall_consecutive_required) + 1))
        stop_reason = "running"

        print("=== 开始夹爪标定数据采集 ===")
        print(f"run_id: {run_id}")
        print(f"direction: {direction}, step_rad: {step_rad}, settle_delay_s: {settle_delay_s}")
        print(f"csv: {csv_path}")

        while True:
            loop_start = time.monotonic()
            elapsed = loop_start - start_t
            if step_index >= max_steps or ((max_duration_s is not None) and float(max_duration_s) > 0 and elapsed >= float(max_duration_s)):
                stop_reason = "timeout"
                break

            current = self.safe_get_feedback_snapshot(pos_name=pos_name)
            current_pos = current["gripper_pos"]
            if current_pos is None:
                stop_reason = "invalid_feedback"
                break

            command_pos = float(current_pos) + direction_sign * float(step_rad)
            self.g.robot.set_actions({part: {"type": "position", "position": [command_pos]}})

            if settle_delay_s > 0:
                # 每步位移后保持1s（默认）让激光和机构稳定
                self._hold_position_command(part=part, target=command_pos, hold_s=settle_delay_s)

            sample = self.safe_get_feedback_snapshot(pos_name=pos_name)
            actual_pos = sample["gripper_pos"]

            delta_pos = None
            if actual_pos is not None:
                pos_history.append(float(actual_pos))
            if (prev_actual_pos is not None) and (actual_pos is not None):
                delta_pos = abs(float(actual_pos) - float(prev_actual_pos))
                delta_history.append(delta_pos)

            row_reason = "running"
            if actual_pos is None:
                row_reason = "invalid_feedback"
                stop_reason = "invalid_feedback"
            elif (step_index >= int(min_steps_before_stall)) and self.detect_mechanical_limit_by_stall(
                delta_history=delta_history,
                pos_history=pos_history,
                threshold=float(stall_delta_threshold),
                consecutive_count=int(stall_consecutive_required),
                window_motion_threshold=float(stall_window_motion_threshold),
            ):
                row_reason = "stall_limit"
                stop_reason = "stall_limit"

            row = {
                "timestamp": self._format_timestamp_ms(),
                "run_id": run_id,
                "gripper_name": part,
                "direction": direction,
                "step_index": step_index,
                "command_pos_rad": command_pos,
                "gripper_pos": actual_pos,
                "gripper_opening_distance": sample["real_distance"],
                "gripper opening distance": sample["real_distance"],
                "distance": sample["real_distance"],
                "delta_pos_from_prev": delta_pos,
                "gripper_vel": sample["gripper_vel"],
                "gripper_torque": sample["gripper_torque"],
                "rb_time": sample["rb_time"],
                "laser_status": sample["laser_status"],
                "limit_reason": row_reason,
            }
            self.append_csv_row(csv_path=csv_path, row_dict=row)

            # 每一步打印当前采样状态，便于在线观察采集质量
            command_str = f"{command_pos:.6f}" if command_pos is not None else "None"
            actual_str = f"{actual_pos:.6f}" if actual_pos is not None else "None"
            laser_val = sample["real_distance"]
            laser_str = f"{laser_val:.3f}" if laser_val is not None else "None"
            delta_str = f"{delta_pos:.6f}" if delta_pos is not None else "None"
            print(
                f"[collect][{direction}] step={step_index:05d}, cmd={command_str}, "
                f"pos={actual_str}, laser={laser_str}, dpos={delta_str}, reason={row_reason}"
            )

            step_index += 1
            prev_actual_pos = actual_pos

            if row_reason != "running":
                break

            if control_interval_s > 0:
                time.sleep(control_interval_s)

        final_sample = self.safe_get_feedback_snapshot(pos_name=pos_name)
        hold_pos = final_sample["gripper_pos"]
        if hold_pos is None:
            hold_pos = prev_actual_pos
        if hold_pos is not None:
            try:
                self.g.robot.set_actions({part: {"type": "position", "position": [float(hold_pos)]}})
            except Exception:
                pass

        print(f"采集结束: reason={stop_reason}, steps={step_index}")
        print(f"数据已写入: {csv_path}")
        self.g.loggerUI.info(
            f"gripper calibration done: run_id={run_id}, direction={direction}, reason={stop_reason}, steps={step_index}, csv={csv_path}"
        )

        return {
            "run_id": run_id,
            "csv_path": csv_path,
            "direction": direction,
            "steps": step_index,
            "limit_reason": stop_reason,
        }

    @hide_ui_while
    def run_open_calibration(self, part: str, pos_name: str = "gripper_pos"):
        return self.collect_gripper_calibration_data(
            part=part,
            pos_name=pos_name,
            direction="open",
        )

    @hide_ui_while
    def run_close_calibration(self, part: str, pos_name: str = "gripper_pos"):
        return self.collect_gripper_calibration_data(
            part=part,
            pos_name=pos_name,
            direction="close",
        )

    def _hold_position_command(self, part: str, target: float, hold_s: float):
        end_t = time.monotonic() + hold_s
        cmd = [float(target)]
        while time.monotonic() < end_t:
            self.g.robot.set_actions({part: {"type": "position", "position": cmd}})
            time.sleep(1 / CONTROL_HZ)

    def _recover_motion_after_zero(self, part: str, pos_name: str = "gripper_pos"):
        # Re-latch position mode and resend a short hold command to avoid
        # post-set_zero non-responsive state seen on some drivers.
        try:
            self.g.robot.send_command(part, {"command": "set_control_mode", "mode": "position"})
        except Exception:
            pass
        time.sleep(0.2)
        cur = self._get_feedback_scalar(pos_name)
        if cur is None:
            cur = 0.0
        self._hold_position_command(part=part, target=cur, hold_s=0.4)

    def _write_gripper_yaml_params(self, part: str, length_per_radian: float = None, offset_at_hardware_zero: float = None):
        yaml_path = f"/opt/robot/rb_hardware/{part}.yaml"
        with open(yaml_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}

        if length_per_radian is not None:
            old = data.get("length_per_radian")
            data["length_per_radian"] = [float(length_per_radian)] if isinstance(old, list) else float(length_per_radian)
        if offset_at_hardware_zero is not None:
            old = data.get("offset_at_hardware_zero")
            data["offset_at_hardware_zero"] = [float(offset_at_hardware_zero)] if isinstance(old, list) else float(offset_at_hardware_zero)

        with open(yaml_path, "w", encoding="utf-8") as f:
            yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)

        return yaml_path

    def _set_following_error_window_value(self, part: str, window_size: int):
        self.g.robot.send_command(
            part,
            {"command": "set_following_error_window", "value": [int(window_size)]},
        )

    def set_torque_mode(self,part):
        confirm = input("请在切换力矩模式前托住机械臂，否则机械臂将自由下落 (y/n): ").strip().lower()
        if confirm == "y":
            print("已切换到力矩模式")
            self.g.robot.send_command(part, {"command": "set_control_mode", "mode": "torque"})
        else:
            print("Cancelled.")
            input("回车以返回")
            return
        confirm = input("输入y以离开力矩模式返回到位置模式 (y)")
        if confirm == "y":
            print("已切换到位置模式")
            self.g.robot.send_command(part, {"command":"set_control_mode", "mode": "position"})
            input("回车以返回")
            return
    def _run_point(self,q_name,q_pos,part,timeout_s=3.0):
        '''
        _run_point 的 Docstring
        
        :param q_name: self.g.feedbackData.QActual for arm
        :param q_pos: target pos
        :param part: self.g.selected_arm for arm
        :param timeout_s: timeout seconds 
        '''
        print("Moving...\n")
        start_time = time.monotonic()
        while True:
            cur_q = getattr(self.g.feedbackData, q_name)
            if np.max(np.abs(cur_q - np.array(q_pos))) <= TOL:
                print("到达！")
                self.g.loggerUI.info("到达！")
                break
            if time.monotonic() - start_time > timeout_s:
                print(f"走点在{timeout_s}后超时退出")
                self.g.loggerUI.warn(f"走点在{timeout_s}后超时退出")
                break
            self.g.robot.set_actions({part: {"type": "position", "position": q_pos}})
            time.sleep(1 / CONTROL_HZ)
        input("回车以结束")

    @hide_ui_while
    def _run_point_with_input(self,q_name,dof,part):
        '''
        _run_point_with_input 的 Docstring
        
        :param cur_q: self.g.feedbackData.QActual for arm
        :param dof: target pos 
        :param part: self.g.selected_arm for arm 
        '''
        try:
            raw_input = input(f"以逗号分隔的方式输入N自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N 或 'q' 退出): ").strip()
            if raw_input.lower() == 'q':
                print("结束")
                return
            q_pos = [float(x.strip()) for x in raw_input.split(',')]

            if len(q_pos) != dof:
                print(f"\nError: Expected {dof} values, but got {len(q_pos)}.")
                return
            raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
            timeout_s = float(raw_input) if raw_input else 3.0
            print(f"\n移动到新位置: {[f'{p:.3f}' for p in q_pos]}")
            self._run_point(q_name=q_name,q_pos=q_pos,part=part,timeout_s=timeout_s)
        
        except Exception as e:
            print(f"\nError:{e}")
            input("回车以返回上一级目录")

    def _run_point_without_hide(self,q_name,dof,part):
        '''
        _run_point_without_hide 的 Docstring
        
        :param cur_q: self.g.feedbackData.QActual for arm
        :param dof: target pos 
        :param part: self.g.selected_arm for arm 
        '''
        try:
            while True:
                raw_input = input(f"以逗号分隔的方式输入{dof}自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N): ")
                try:
                    q_pos = [float(x.strip()) for x in raw_input.split(',')]
                except ValueError:
                    print("\nError: 输入中包含无法转换为数字的值，请重新输入。")
                    continue
                if len(q_pos) != dof:
                    print(f"\nError: 期望 {dof} 个值，但输入了 {len(q_pos)} 个，请重新输入。")
                    continue
                # 如果长度正确，退出循环
                break
            raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
            timeout_s = float(raw_input) if raw_input else 3.0
            print(f"\n移动到新位置: {[f'{p:.3f}' for p in q_pos]}")
            self._run_point(q_name=q_name,q_pos=q_pos,part=part,timeout_s=timeout_s)
        
        except Exception as e:
            print(f"\nError:{e}")
            input("回车以返回上一级目录")

    @hide_ui_while
    def set_max_vel_acc(self):
        os.system('clear')
        # Helper 函数：安全读取正浮点数
        def read_positive_float(prompt):
            while True:
                raw_input = input(prompt)
                try:
                    value = float(raw_input)
                    if value <= 0:
                        print("⚠️ 必须输入正数，请重试。")
                        continue
                    return value
                except ValueError:
                    print("⚠️ 输入无效，请输入数字。")
        # 读取最大速度
        self.max_vel = read_positive_float("请输入最大速度设置 (rad/s): ")
        # 读取最大加速度
        self.max_acc = read_positive_float("请输入最大加速度设置 (rad/s²): ")
        input(f"✅ 设置完成：最大速度 = {self.max_vel} rad/s, 最大加速度 = {self.max_acc} rad/s²\n回车以继续...")

    @hide_ui_while
    def cubic_move(self,start_data,max_vel:list,max_acc:list,part:str,pos_name:str,dof:int):
        '''
        cubic_move_record 的 Docstring

        :param start_data: self.g.feedbackData.QActual
        :param max_vel: [self.max_vel]*7 for arm
        :type max_vel: list
        :param max_acc: [self.max_acc]*7 for arm
        :type max_acc: list
        :part: self.g.selected_arm
        :type part: str
        :pos_name: getattr(self.g.feedbackData, "rb_time")
        :type pos_name: str
        '''
        start_point = np.array(start_data,dtype=float)
        
        print("当前点位:")
        print(" ".join([f"{v:.6f}" for v in start_point]))
        
        raw_input = input(f"以逗号分隔的方式输入N自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N 或 'q' 退出): ")
        if raw_input.lower() == 'q':
            print("结束")
            return
        end_point = [float(x.strip()) for x in raw_input.split(',')]

        if len(end_point) != dof:
            self.g.loggerUI.error(f"Expected {dof} values, but got {len(end_point)}.")
            return

        print("\n最终目标点位:")
        print(end_point)
        input("确认目标点位，回车以开始走点")
        # 构造路径
        path = np.vstack([start_point,end_point])

        # 轨迹规划
        sample_rate = 20
        step = 1/sample_rate
        try:
            ts, qs, qds, qdds, duration = TOPP(
                path, 
                max_vel, 
                max_acc, 
                step
            )
        except RuntimeError as e:
            print("轨迹规划失败:", e)
            return

        # 执行轨迹
        print("移动中...")
        
        start_time = time.time()
        for t, q in zip(ts, qs):
            self.g.robot.set_actions({part: {"type": "position", "position": q.tolist()}})
            while time.time() - start_time < t:
                time.sleep(0.05)
        print("移动结束")

    @hide_ui_while
    def cubic_move_record(self,start_data,max_vel:list,max_acc:list,part:str,pos_name:str,dof:int):
        '''
        cubic_move_record 的 Docstring

        :param start_data: self.g.feedbackData.QActual
        :param max_vel: [self.max_vel]*7 for arm
        :type max_vel: list
        :param max_acc: [self.max_acc]*7 for arm
        :type max_acc: list
        :part: self.g.selected_arm
        :type part: str
        :pos_name: getattr(self.g.feedbackData, "rb_time")
        :type pos_name: str
        '''
        start_point = np.array(start_data,dtype=float)
        # 执行轨迹前清空日志
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()
        # 减少反馈负担
        self.g.feedback_mode = self.g.feedback_mode.BASIC
        print("当前点位:")
        print(" ".join([f"{v:.6f}" for v in start_point]))
        
        raw_input = input(f"以逗号分隔的方式输入N自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N): ")
        end_point = [float(x.strip()) for x in raw_input.split(',')]
        if len(end_point) != dof:
            self.g.loggerUI.error(f"Expected {dof} values, but got {len(end_point)}.")
            return

        print("\n最终目标点位:")
        print(end_point)
        input("确认目标点位，回车以开始走点")
        # 构造路径
        path = np.vstack([start_point,end_point])

        # 轨迹规划
        sample_rate = 20
        step = 1/sample_rate
        try:
            ts, qs, qds, qdds, duration = TOPP(
                path, 
                max_vel, 
                max_acc, 
                step
            )
        except RuntimeError as e:
            print("轨迹规划失败:", e)
            return

        # 执行轨迹
        print("移动中...")
        self.g.record_flag.set()
        print("开始高频+低频记录...")
        start_time = time.time()
        for t, q in zip(ts, qs):
            # 使用锁保护读取操作
            with self.g.feedback_lock:
                cur_q = getattr(self.g.feedbackData, pos_name)
                rb_time = getattr(self.g.feedbackData, "rb_time")
            self.g.robot.set_actions({part: {"type": "position", "position": q.tolist()}})
            self.g.logger._record_lowfreq(t=time.perf_counter(),q=q,
                                          rb_time=rb_time,
                                          feedback_pos=cur_q)
            while time.time() - start_time < t:
                time.sleep(0.05)
        # 延迟结束
        extra_record = time.perf_counter()
        while time.perf_counter() - extra_record <= EXTRA_TIME:
            with self.g.feedback_lock:
                cur_q = getattr(self.g.feedbackData, pos_name)
                rb_time = getattr(self.g.feedbackData, "rb_time")
            self.g.logger._record_lowfreq(t=time.time(),q=q,
                                    rb_time=rb_time,
                                    feedback_pos=cur_q)
            time.sleep(1 / CONTROL_HZ)
        self.g.record_flag.clear()
        print("高频记录结束，共收集高频数据", len(self.g.highfreq_log))
        print("低频记录结束，共收集低频数据", len(self.g.lowfreq_log))
        # 减少反馈负担
        self.g.feedback_mode = self.g.feedback_mode.FULL
        print("\n=== 数据记录 ===")
        confirm = input("是否保存日志? (y/n): ").strip().lower()
        if confirm != "y":
            print("日志保存更新取消")
            input("回车以返回")
            return

        # 保存日志
        highfreq_filename, lowfreq_filename,tstamp = self.g.logger._save_logs(part)

        confirm = input("是否可视化记录数据? (y/n): ").strip().lower()
        if confirm != "y":
            print("数据可视化取消")
            input("回车以返回")
            return

        draw(log_dir=f'{project_root}/logs',lowfile=lowfreq_filename,highfile=highfreq_filename,savefig=f"viz_{tstamp}.png")

    @hide_ui_while
    def step_move_record(self,part:str,pos_name:str):
        '''
        step_move_record 的 Docstring

        :part: self.g.selected_arm
        :type part: str
        :pos_name: getattr(self.g.feedbackData, "rb_time")
        :type pos_name: str
        '''
        # 执行轨迹前清空日志
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()
        # 减少反馈负担
        self.g.feedback_mode = self.g.feedback_mode.BASIC
        user_input = input("请输入目标位置 (逗号分隔) 或 'q' 退出: ").strip()
        if user_input.lower() == 'q':
            print("结束单步移动记录")
            return
        raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
        timeout_s = float(raw_input) if raw_input else 3.0
        try:
            q_pos = [float(x.strip()) for x in user_input.split(',')]
            print(f"移动到新位置: {[f'{p:.3f}' for p in q_pos]}")
            
            # 执行移动并记录
            self.g.record_flag.set()
            # 占位
            print("开始高频+低频记录...")
            print("Moving...\n")
            start_time = time.monotonic()
            while True:
                # 使用锁保护读取操作
                with self.g.feedback_lock:
                    cur_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                if np.max(np.abs(cur_q - np.array(q_pos))) <= TOL:
                    print("到达！")
                    self.g.loggerUI.info("到达！")
                    break
                elif time.monotonic() - start_time > timeout_s:
                    print(f"走点在{timeout_s}后超时退出")
                    self.g.loggerUI.warn(f"走点在{timeout_s}后超时退出")
                    break
                self.g.robot.set_actions({part: {"type": "position", "position": q_pos}})
                self.g.logger._record_lowfreq(t=time.perf_counter(),q=q_pos,
                                        rb_time=rb_time,
                                        feedback_pos=cur_q)
                time.sleep(1 / CONTROL_HZ)
            # 延迟结束
            extra_record = time.perf_counter()
            while time.perf_counter() - extra_record <= EXTRA_TIME:
                with self.g.feedback_lock:
                    cur_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                self.g.logger._record_lowfreq(t=time.time(),q=q_pos,
                                        rb_time=rb_time,
                                        feedback_pos=cur_q)
                time.sleep(1 / CONTROL_HZ)
            self.g.record_flag.clear()
            print("高频记录结束，共收集高频数据", len(self.g.highfreq_log))
            print("低频记录结束，共收集低频数据", len(self.g.lowfreq_log))
            # 减少反馈负担
            self.g.feedback_mode = self.g.feedback_mode.FULL
            print("\n=== 数据记录 ===")
            confirm = input("是否保存日志? (y/n): ").strip().lower()
            if confirm != "y":
                print("日志保存更新取消")
                input("回车以返回")
                return

            # 保存日志
            highfreq_filename, lowfreq_filename,tstamp = self.g.logger._save_logs(part)

            confirm = input("是否可视化记录数据? (y/n): ").strip().lower()
            if confirm != "y":
                print("数据可视化取消")
                input("回车以返回")
                return

            # draw(log_dir=f'{project_root}/logs',lowfile=lowfreq_filename,highfile=highfreq_filename,savefig=f"viz_{tstamp}.png")
            draw_step_response_analysis(log_dir=f'{project_root}/logs',
                 lowfile=lowfreq_filename,
                 highfile=highfreq_filename,
                 savefig=f"viz_{tstamp}.png",
                 target_pos=q_pos,
                 threshold=0.02,
                 )

        except Exception as e:
            print(e)

    def _run_step_record_once_auto(
        self,
        part: str,
        pos_name: str,
        target_pos: list,
        timeout_s: float = 3.0,
        threshold: float = 0.02,
    ):
        """Run one step response record without interactive prompts and return artifact paths."""
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()
        self.g.feedback_mode = self.g.feedback_mode.BASIC

        q_pos = [float(x) for x in target_pos]
        reached = False
        timed_out = False

        try:
            self.g.record_flag.set()
            print(f"[自动阶跃] 开始记录，目标位置: {q_pos}")

            start_time = time.monotonic()
            while True:
                with self.g.feedback_lock:
                    cur_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")

                if np.max(np.abs(np.array(cur_q) - np.array(q_pos))) <= TOL:
                    reached = True
                    print("[自动阶跃] 到达目标")
                    self.g.loggerUI.info("[自动阶跃] 到达目标")
                    break

                if time.monotonic() - start_time > timeout_s:
                    timed_out = True
                    print(f"[自动阶跃] 超时退出: {timeout_s}s")
                    self.g.loggerUI.warn(f"[自动阶跃] 超时退出: {timeout_s}s")
                    break

                self.g.robot.set_actions({part: {"type": "position", "position": q_pos}})
                self.g.logger._record_lowfreq(
                    t=time.perf_counter(),
                    q=q_pos,
                    rb_time=rb_time,
                    feedback_pos=cur_q,
                )
                time.sleep(1 / CONTROL_HZ)

            extra_record = time.perf_counter()
            while time.perf_counter() - extra_record <= EXTRA_TIME:
                with self.g.feedback_lock:
                    cur_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                self.g.logger._record_lowfreq(
                    t=time.time(),
                    q=q_pos,
                    rb_time=rb_time,
                    feedback_pos=cur_q,
                )
                time.sleep(1 / CONTROL_HZ)

            highfreq_filename, lowfreq_filename, tstamp = self.g.logger._save_logs(part)
            highfreq_filename = self._move_log_file_to_output_dir(highfreq_filename)
            lowfreq_filename = self._move_log_file_to_output_dir(lowfreq_filename)
            fig_name = f"viz_{tstamp}.png"
            out_dir = self._get_output_dir()
            draw_step_response_analysis(
                log_dir=out_dir,
                lowfile=lowfreq_filename,
                highfile=highfreq_filename,
                savefig=fig_name,
                target_pos=q_pos,
                threshold=threshold,
                show_plot=False,
            )

            return {
                "target_pos": q_pos,
                "reached": reached,
                "timed_out": timed_out,
                "highfreq_file": highfreq_filename,
                "lowfreq_file": lowfreq_filename,
                "plot_path": os.path.join(out_dir, fig_name),
                "vel_plot_path": os.path.join(out_dir, f"vel_{fig_name}"),
                "temp_plot_path": os.path.join(out_dir, f"temp_{fig_name}"),
            }
        finally:
            self.g.record_flag.clear()
            self.g.feedback_mode = self.g.feedback_mode.FULL

    def _write_step_response_report_xlsx(
        self,
        rows: list,
        report_prefix: str = "step_response_report",
        second_col_title: str = "位置数据可视化",
    ):
        """
        Write a new xlsx report and embed images.
        columns: 目标位置 / 位置数据可视化 / 速度数据可视化
        """
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise RuntimeError(
                f"缺少依赖 openpyxl。当前解释器: {sys.executable}，"
                "请在该解释器对应环境安装: pip install openpyxl"
            ) from e

        image_embed_enabled = True
        image_import_error = None
        try:
            from openpyxl.drawing.image import Image as XLImage
        except Exception as e:
            XLImage = None
            image_embed_enabled = False
            image_import_error = e

        report_ts = time.strftime("%Y%m%d_%H%M%S")
        out_dir = self._get_output_dir()
        report_path = os.path.join(out_dir, f"{report_prefix}_{report_ts}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "step_response"

        ws["A1"] = "目标位置"
        ws["B1"] = second_col_title
        ws["C1"] = "速度数据可视化"
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 56
        ws.column_dimensions["C"].width = 56

        if not image_embed_enabled:
            warn_msg = (
                "图片嵌入不可用（通常是缺少 Pillow），当前降级为写入图片路径。"
                f" 错误: {image_import_error}"
            )
            print(warn_msg)
            self.g.loggerUI.warn(warn_msg)

        for idx, row in enumerate(rows, start=2):
            ws.cell(row=idx, column=1, value=row["target_label"])
            ws.row_dimensions[idx].height = 180

            pos_img_path = row["plot_path"]
            vel_img_path = row["vel_plot_path"]
            pos_img_embedded = False
            vel_img_embedded = False

            if image_embed_enabled and os.path.exists(pos_img_path):
                try:
                    pos_img = XLImage(pos_img_path)
                    pos_img.width = 380
                    pos_img.height = 210
                    pos_img.anchor = f"B{idx}"
                    ws.add_image(pos_img)
                    pos_img_embedded = True
                except Exception as e:
                    ws.cell(row=idx, column=2, value=f"插图失败，路径: {pos_img_path}, err: {e}")
            else:
                ws.cell(row=idx, column=2, value=pos_img_path if os.path.exists(pos_img_path) else f"图片不存在: {pos_img_path}")

            if image_embed_enabled and os.path.exists(vel_img_path):
                try:
                    vel_img = XLImage(vel_img_path)
                    vel_img.width = 380
                    vel_img.height = 210
                    vel_img.anchor = f"C{idx}"
                    ws.add_image(vel_img)
                    vel_img_embedded = True
                except Exception as e:
                    ws.cell(row=idx, column=3, value=f"插图失败，路径: {vel_img_path}, err: {e}")
            else:
                ws.cell(row=idx, column=3, value=vel_img_path if os.path.exists(vel_img_path) else f"图片不存在: {vel_img_path}")

            row["_report_images_embedded"] = pos_img_embedded and vel_img_embedded

        wb.save(report_path)
        return report_path

    def _cleanup_report_artifacts(self, rows: list):
        log_dir = self._get_output_dir()
        for row in rows:
            for key in ("highfreq_file", "lowfreq_file"):
                log_path = row.get(key)
                if not log_path:
                    continue
                if not os.path.isabs(log_path):
                    log_path = os.path.join(log_dir, log_path)
                try:
                    if os.path.exists(log_path):
                        os.remove(log_path)
                except OSError as e:
                    self.g.loggerUI.warn(f"临时数据删除失败: {log_path}, err: {e}")

            for key in ("plot_path", "vel_plot_path", "temp_plot_path"):
                image_path = row.get(key)
                if not image_path:
                    continue
                try:
                    if os.path.exists(image_path):
                        os.remove(image_path)
                except OSError as e:
                    self.g.loggerUI.warn(f"临时图片删除失败: {image_path}, err: {e}")

    def _cleanup_auto_output_artifacts(self, out_dir: str):
        if not out_dir or not os.path.isdir(out_dir):
            return
        transient_keywords = ("viz", "tmp", "temp", "highfreq", "lowfreq")
        for path in Path(out_dir).iterdir():
            if not path.is_file():
                continue
            name = path.name.lower()
            should_delete = path.suffix.lower() == ".png" or (
                path.suffix.lower() == ".json" and any(key in name for key in transient_keywords)
            )
            if not should_delete:
                continue
            try:
                path.unlink()
            except OSError as e:
                self.g.loggerUI.warn(f"临时文件删除失败: {path}, err: {e}")

    def _run_topp_record_once_auto(
        self,
        part: str,
        pos_name: str,
        target_pos: list,
        timeout_s: float = 3.0,
    ):
        """
        Run one TOPP trajectory tracking record without interactive prompts.
        """
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()
        self.g.feedback_mode = self.g.feedback_mode.BASIC

        try:
            with self.g.feedback_lock:
                cur_q = getattr(self.g.feedbackData, pos_name)
            start_point = np.array(cur_q, dtype=float)
            end_point = np.array(target_pos, dtype=float)
            dof = len(end_point)
            path = np.vstack([start_point, end_point])

            sample_rate = 20
            step = 1 / sample_rate
            max_vel = [float(self.max_vel)] * dof
            max_acc = [float(self.max_acc)] * dof
            try:
                ts, qs, qds, qdds, duration = TOPP(path, max_vel, max_acc, step)
            except RuntimeError as e:
                self.g.loggerUI.error(f"[TOPP自动化] 轨迹规划失败: {e}")
                raise

            self.g.record_flag.set()
            print(f"[TOPP自动化] 开始记录，目标位置: {end_point.tolist()}")
            start_time = time.time()
            for t, q in zip(ts, qs):
                with self.g.feedback_lock:
                    fb_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                self.g.robot.set_actions({part: {"type": "position", "position": q.tolist()}})
                self.g.logger._record_lowfreq(
                    t=time.perf_counter(),
                    q=q,
                    rb_time=rb_time,
                    feedback_pos=fb_q,
                )
                while time.time() - start_time < t:
                    time.sleep(0.05)
                if time.time() - start_time > timeout_s and t < ts[-1]:
                    self.g.loggerUI.warn(f"[TOPP自动化] 执行超时 {timeout_s}s，提前结束")
                    break

            extra_record = time.perf_counter()
            while time.perf_counter() - extra_record <= EXTRA_TIME:
                with self.g.feedback_lock:
                    fb_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                self.g.logger._record_lowfreq(
                    t=time.time(),
                    q=end_point,
                    rb_time=rb_time,
                    feedback_pos=fb_q,
                )
                time.sleep(1 / CONTROL_HZ)

            highfreq_filename, lowfreq_filename, tstamp = self.g.logger._save_logs(part)
            highfreq_filename = self._move_log_file_to_output_dir(highfreq_filename)
            lowfreq_filename = self._move_log_file_to_output_dir(lowfreq_filename)
            fig_name = f"viz_{tstamp}.png"
            out_dir = self._get_output_dir()
            draw(
                log_dir=out_dir,
                lowfile=lowfreq_filename,
                highfile=highfreq_filename,
                savefig=fig_name,
                show_plot=False,
            )

            return {
                "target_pos": end_point.tolist(),
                "highfreq_file": highfreq_filename,
                "lowfreq_file": lowfreq_filename,
                "plot_path": os.path.join(out_dir, fig_name),
                "vel_plot_path": os.path.join(out_dir, f"vel_{fig_name}"),
                "temp_plot_path": os.path.join(out_dir, f"temp_{fig_name}"),
            }
        finally:
            self.g.record_flag.clear()
            self.g.feedback_mode = self.g.feedback_mode.FULL

    @hide_ui_while
    def step_response_auto_test(self, part: str, pos_name: str):
        """
        基础功能测试-阶跃响应测试（自动化）:
        两轮往返，共4条记录：
        1) 0 -> 0.05
        2) 0.05 -> 0
        3) 0 -> 0.05
        4) 0.05 -> 0
        """
        rounds = 2
        settle_wait_s = 3.0
        timeout_s = 3.0
        threshold = 0.02
        up_target = 0.05
        down_target = 0.0
        rows = []

        print("=== 自动阶跃响应测试开始 ===")

        for r in range(rounds):
            print(f"[Round {r+1}/{rounds}] 先归零到 0.0")
            zero_ok = self._move_to_target(part=part, pos_name=pos_name, target=0.0, timeout_s=timeout_s)
            if not zero_ok:
                self.g.loggerUI.warn(f"[Round {r+1}] 归零超时，继续执行后续测试")
                print(f"[Round {r+1}] 归零超时，继续执行后续测试")
            time.sleep(settle_wait_s)

            up = self._run_step_record_once_auto(
                part=part,
                pos_name=pos_name,
                target_pos=[up_target],
                timeout_s=timeout_s,
                threshold=threshold,
            )
            up["target_label"] = "0->0.05"
            rows.append(up)
            time.sleep(settle_wait_s)

            down = self._run_step_record_once_auto(
                part=part,
                pos_name=pos_name,
                target_pos=[down_target],
                timeout_s=timeout_s,
                threshold=threshold,
            )
            down["target_label"] = "0.05->0"
            rows.append(down)
            time.sleep(settle_wait_s)

        report_path = self._write_step_response_report_xlsx(rows)
        self._cleanup_report_artifacts(rows)
        print(f"=== 自动阶跃响应测试结束，报告已生成: {report_path} ===")
        self.g.loggerUI.info("阶跃响应测试完成")

    @hide_ui_while
    def topp_tracking_auto_test(self, part: str, pos_name: str):
        """
        基础功能测试-轨迹跟踪测试（自动化）:
        两轮往返，共4条记录：
        1) 0 -> 0.05
        2) 0.05 -> 0
        3) 0 -> 0.05
        4) 0.05 -> 0
        固定 TOPP 参数:
        - max_vel = 0.1 rad/s
        - max_acc = 1.0 rad/s^2
        """
        rounds = 2
        settle_wait_s = 3.0
        timeout_s = 3.0
        up_target = 0.05
        down_target = 0.0
        rows = []

        self.max_vel = 0.1
        self.max_acc = 1.0
        print("=== 自动轨迹跟踪测试开始 ===")
        print(f"固定TOPP参数: max_vel={self.max_vel} rad/s, max_acc={self.max_acc} rad/s^2")

        for r in range(rounds):
            print(f"[Round {r+1}/{rounds}] 先归零到 0.0")
            zero_ok = self._move_to_target(part=part, pos_name=pos_name, target=0.0, timeout_s=timeout_s)
            if not zero_ok:
                self.g.loggerUI.warn(f"[Round {r+1}] 归零超时，继续执行后续测试")
                print(f"[Round {r+1}] 归零超时，继续执行后续测试")
            time.sleep(settle_wait_s)

            up = self._run_topp_record_once_auto(
                part=part,
                pos_name=pos_name,
                target_pos=[up_target],
                timeout_s=timeout_s,
            )
            up["target_label"] = "0->0.05"
            rows.append(up)
            time.sleep(settle_wait_s)

            down = self._run_topp_record_once_auto(
                part=part,
                pos_name=pos_name,
                target_pos=[down_target],
                timeout_s=timeout_s,
            )
            down["target_label"] = "0.05->0"
            rows.append(down)
            time.sleep(settle_wait_s)

        report_path = self._write_step_response_report_xlsx(
            rows=rows,
            report_prefix="topp_tracking_report",
            second_col_title="位置及跟踪误差数据可视化",
        )
        self._cleanup_report_artifacts(rows)
        print(f"=== 自动轨迹跟踪测试结束，报告已生成: {report_path} ===")
        self.g.loggerUI.info("轨迹跟踪测试完成")

    def basic_function_auto_test(self, part: str, pos_name: str):
        ui = getattr(self.g, "ui", None)
        hidden_for_auto = False
        if ui is not None and getattr(ui, "show_ui", False):
            ui.simulate_key("h")
            hidden_for_auto = True
        gripper_no = input("请输入夹爪 ID: ").strip()
        if not gripper_no:
            gripper_no = "unknown"
        safe_no = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in gripper_no)
        tstamp = time.strftime("%Y%m%d_%H%M%S")
        prev_output_dir = self.current_output_dir
        sn_label = getattr(self.g, "gripper_serial_label", "SN_unknown") or "SN_unknown"
        sn_label = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(sn_label))
        self.current_output_dir = os.path.join(project_root, "logs", sn_label, f"{safe_no}_{tstamp}")
        os.makedirs(self.current_output_dir, exist_ok=True)

        tests = [
            ("零位行程和限位测试", lambda: self.zero_limit_travel_test(part=part, pos_name=pos_name, wait_for_return=False)),
            ("重复定位精度", lambda: self.repeatability_position_accuracy_test(part=part, pos_name=pos_name, wait_for_return=False)),
            ("绝对定位精度", lambda: self.absolute_position_accuracy_test(part=part, pos_name=pos_name, wait_for_return=False)),
            ("一维力精度", lambda: self.one_dim_force_accuracy_test(part=self.g.selected_loadcell, wait_for_return=False)),
            ("阶跃响应测试", lambda: self.step_response_auto_test(part=part, pos_name=pos_name)),
            ("轨迹跟踪测试", lambda: self.topp_tracking_auto_test(part=part, pos_name=pos_name)),
        ]

        print("=== 基础功能自动测试开始 ===")
        print(f"电机SN目录: {sn_label}")
        print(f"结果目录: {self.current_output_dir}")
        self.g.loggerUI.info(f"基础功能自动测试开始 gripper={safe_no} sn={sn_label}")
        success = False
        try:
            for idx, (name, run_test) in enumerate(tests, start=1):
                print(f"\n[{idx}/{len(tests)}] 开始{name}")
                self.g.loggerUI.info(f"[{idx}/{len(tests)}] {name}")
                run_test()
                print(f"[{idx}/{len(tests)}] 完成{name}")
            success = True
        finally:
            out_dir = self.current_output_dir
            self.current_output_dir = prev_output_dir
            if success:
                print("=== 基础功能自动测试结束 ===")
                self.g.loggerUI.info(f"基础功能自动测试完成 gripper={safe_no} dir={os.path.basename(out_dir)}")
            self._cleanup_auto_output_artifacts(out_dir)
            self._clear_hblog_view()
            if hidden_for_auto and ui is not None:
                ui.simulate_key("\n")

    def _clear_hblog_view(self):
        try:
            self.g.hblog_buffer.consume()
        except Exception:
            pass
        ui = getattr(self.g, "ui", None)
        if ui is None:
            return
        try:
            ui._hblog_lines.clear()
            ui.hblog_view_start = 0
            ui.hblog_follow = True
        except Exception:
            pass

    @hide_ui_while
    def set_limits(self,part):
        min_pos = input("输入最小软限位:")
        max_pos = input("输入最大软限位:")
        min_list = [float(min_pos)]
        max_list = [float(max_pos)]
        self.g.robot.send_command(part, {"command": "set_limit",
                                        "enabled": [True],
                                        "lower":min_list,
                                        "upper":max_list})
        self.g.loggerUI.info("已设置软限位为: "+str(min_pos)+" ~ "+str(max_pos))

    @hide_ui_while
    def set_following_error_window(self,part):
        window_size = input("输入电机跟踪误差窗口大小/脉冲整数:")
        window_list = [int(window_size)]
        self.g.robot.send_command(part, {"command": "set_following_error_window",
                                        "value":window_list})
        self.g.loggerUI.info("已设置电机跟踪误差窗口为: "+str(window_size)+" pulses")

    @hide_ui_while
    def set_following_error_window_10000000(self, part: str):
        self._set_following_error_window_value(part=part, window_size=10000000)
        self.g.loggerUI.info("已设置跟踪误差窗口: 10000000")
        print("已设置跟踪误差窗口: 10000000")
        input("回车返回")

    @hide_ui_while
    def manual_calibration_reset_yaml_and_reload(self, part: str, pos_name: str = "gripper_pos"):
        print("Step 1: reset yaml params to length_per_radian=1.0, offset_at_hardware_zero=0.0")
        try:
            yaml_path = self._write_gripper_yaml_params(
                part=part,
                length_per_radian=1.0,
                offset_at_hardware_zero=0.0,
            )
            self.g.loggerUI.info(f"yaml reset done: {yaml_path}")
            print(f"yaml reset done: {yaml_path}")
        except Exception as e:
            self.g.loggerUI.error(f"yaml reset failed: {e}")
            print(f"yaml reset failed: {e}")
            input("Press Enter to return")
            return

        if hasattr(self.g, "reload_robot_from_yaml"):
            try:
                self.g.reload_robot_from_yaml()
                self.g.loggerUI.info("robot reload after yaml reset done")
                print("robot reload done")
            except Exception as e:
                self.g.loggerUI.error(f"robot reload failed: {e}")
                print(f"robot reload failed: {e}")
                input("Press Enter to return")
                return
        else:
            self.g.loggerUI.warn("reload_robot_from_yaml() not found, restart guide manually")
            print("reload function not found, please restart guide manually")

        input("Press Enter to return")

    def _get_gripper_yaml_path(self, part: str) -> str:
        """获取夹爪yaml文件路径"""
        return f"/opt/robot/rb_hardware/{part}.yaml"

    @hide_ui_while
    def sync_yaml_params_to_ui(self, part: str, pos_name: str = "gripper_pos"):
        """同步当前yaml文件的参数到UI（重新加载yaml并下发到电机）"""
        print("=== 同步yaml参数到UI ===")
        print(f"夹爪: {part}")
        
        # 读取当前yaml文件的参数用于显示
        try:
            yaml_path = self._get_gripper_yaml_path(part)
            with open(yaml_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            
            length_per_radian = config.get('length_per_radian', 1.0)
            offset_at_hardware_zero = config.get('offset_at_hardware_zero', 0.0)
            
            print(f"读取到yaml参数:")
            print(f"  length_per_radian = {length_per_radian}")
            print(f"  offset_at_hardware_zero = {offset_at_hardware_zero}")
        except Exception as e:
            print(f"读取yaml文件失败: {e}")
            self.g.loggerUI.error(f"读取yaml文件失败: {e}")
            input("按回车返回")
            return
        
        # 使用reload_robot_from_yaml重新加载yaml并下发到电机
        if hasattr(self.g, "reload_robot_from_yaml"):
            try:
                self.g.reload_robot_from_yaml()
                self.g.loggerUI.info(f"yaml参数同步到UI成功: {part}, length_per_radian={length_per_radian}, offset={offset_at_hardware_zero}")
                print("robot reload done, 参数已下发到电机")
            except Exception as e:
                self.g.loggerUI.error(f"robot reload failed: {e}")
                print(f"robot reload failed: {e}")
                input("按回车返回")
                return
        else:
            self.g.loggerUI.warn("reload_robot_from_yaml() not found, restart guide manually")
            print("reload function not found, please restart guide manually")
        
        print("\n参数同步完成")
        input("按回车返回")
    
    def _calibrate_gripper_climb(self, part: str, pos_name: str, direction_sign: int, 
                                  target_speed: float = 1.0, ctrl_freq: int = 100,
                                  pos_threshold: float = 0.00001, stable_target: int = 15,
                                  use_laser_stop: bool = True) -> float:
        """
        通过位置爬坡模拟匀速运动进行标定
        在接近极限位置时自动减小步长以提高精度
        
        Args:
            part: 夹爪部件名
            pos_name: 位置反馈数据名
            direction_sign: -1 (向内闭合), 1 (向外打开)
            target_speed: 目标速度 (单位/秒)
            ctrl_freq: 控制频率 (Hz)
            pos_threshold: 判断堵转的位置变化阈值
            stable_target: 连续多少次位置不变认为堵转（仅use_laser_stop=False时有效）
            use_laser_stop: True=使用激光距离停止（95mm/5mm），False=使用堵转检测停止
            
        Returns:
            最终位置值
        """
        dt = 1.0 / ctrl_freq
        base_step = target_speed * dt
        current_step = base_step
        
        direction_str = "闭合" if direction_sign == -1 else "张开"
        print(f"\n>>> 开始标定{direction_str}位置 (速度: {target_speed}, 方向: {direction_sign})")
        print(f"    控制频率: {ctrl_freq}Hz, 基础步长: {base_step:.6f}, 堵转阈值: {pos_threshold}")
        print(f"    特性: 接近极限位置时自动减小步长以提高精度")
        
        # 获取起始位置
        start_pos = self._get_feedback_scalar(pos_name)
        if start_pos is None:
            start_pos = 0.0
        command_pos = start_pos
        last_actual_pos = start_pos
        stable_count = 0
        
        # 激光距离阈值（用于调整步长）- 注意：激光数据单位是mm
        LASER_THRESHOLD_OPEN = 90.0    # 张开时激光距离大于90mm时减速（接近100mm的极限）
        LASER_THRESHOLD_CLOSE = 5.0    # 闭合时激光距离小于5mm时减速（接近0mm的极限）
        STEP_REDUCTION_FACTOR = 0.05   # 减速因子（步长变为原来的5%，更精细）
        
        # 基于激光距离的停止阈值
        LASER_STOP_OPEN = 95.0         # 张开时激光距离达到95mm停止
        LASER_STOP_CLOSE = 5.0         # 闭合时激光距离达到5mm停止
        
        # 位置稳定检测参数
        POS_STABLE_THRESHOLD = 0.0001  # 位置变化小于此值认为稳定
        POS_STABLE_COUNT = 50          # 连续50次（0.5秒）稳定认为真正停止
        MAX_WAIT_TIME = 5.0            # 最大等待时间5秒
        
        print(f"    起始位置: {start_pos:.6f}")
        print(f"    停止模式: {'激光停止' if use_laser_stop else '堵转检测'}")
        print("    开始位置爬坡... (按Ctrl+C可中断)")
        
        last_print_time = time.time()
        start_time = time.time()
        slowdown_active = False
        target_reached = False
        target_reached_time = None
        pos_stable_count = 0           # 位置稳定计数
        last_stable_pos = None         # 上次检查的位置
        
        # 堵转检测计数（仅use_laser_stop=False时使用）
        stall_count = 0
        
        try:
            while True:
                # 如果已经到达目标位置，不再下发任何指令，只等待稳定
                if target_reached:
                    # 只读取位置和激光数据，用于显示
                    current_actual_pos = self._get_feedback_scalar(pos_name)
                    if current_actual_pos is None:
                        current_actual_pos = last_actual_pos
                    real_distance = self._get_feedback_scalar("real_distance")
                    
                    # 持续发送当前位置作为目标，确保电机停止运动
                    self.g.robot.set_actions({part: {"type": "position", "position": [current_actual_pos]}})
                    
                    # 检查位置是否稳定
                    if last_stable_pos is not None:
                        pos_change = abs(current_actual_pos - last_stable_pos)
                        if pos_change < POS_STABLE_THRESHOLD:
                            pos_stable_count += 1
                        else:
                            pos_stable_count = 0
                    last_stable_pos = current_actual_pos
                    
                    time.sleep(dt)
                else:
                    # 获取激光测距仪数据（用于调整步长）
                    real_distance = self._get_feedback_scalar("real_distance")
                    
                    # 根据激光距离调整步长（两种模式都使用）
                    if real_distance is not None:
                        if direction_sign == 1:  # 张开 - 激光值越大表示越张开
                            if real_distance > LASER_THRESHOLD_OPEN:  # 接近100mm极限
                                if not slowdown_active:
                                    current_step = base_step * STEP_REDUCTION_FACTOR
                                    slowdown_active = True
                                    print(f"    ↓ 接近张开极限，步长减小至 {current_step:.6f} (激光: {real_distance:.1f}mm)")
                            else:
                                if slowdown_active:
                                    current_step = base_step
                                    slowdown_active = False
                                    print(f"    ↑ 恢复正常步长 {current_step:.6f}")
                        else:  # 闭合 - 激光值越小表示越闭合
                            if real_distance < LASER_THRESHOLD_CLOSE:  # 接近0mm极限
                                if not slowdown_active:
                                    current_step = base_step * STEP_REDUCTION_FACTOR
                                    slowdown_active = True
                                    print(f"    ↓ 接近闭合极限，步长减小至 {current_step:.6f} (激光: {real_distance:.1f}mm)")
                            else:
                                if slowdown_active:
                                    current_step = base_step
                                    slowdown_active = False
                                    print(f"    ↑ 恢复正常步长 {current_step:.6f}")
                    
                    # 1. 更新指令位置
                    command_pos += current_step * direction_sign
                    
                    # 2. 发送位置指令
                    self.g.robot.set_actions({part: {"type": "position", "position": [command_pos]}})
                    
                    # 3. 等待一个周期
                    time.sleep(dt)
                    
                    # 4. 获取当前实际位置
                    current_actual_pos = self._get_feedback_scalar(pos_name)
                    if current_actual_pos is None:
                        current_actual_pos = last_actual_pos
                    
                    # 5. 停止检测
                    if use_laser_stop:
                        # 基于激光距离的停止检测（阶段1使用）
                        if real_distance is not None:
                            if direction_sign == 1:  # 张开
                                if real_distance >= LASER_STOP_OPEN:
                                    target_reached = True
                                    target_reached_time = time.time()
                                    last_stable_pos = current_actual_pos
                                    print(f"\n    ✓ 到达张开目标位置 (激光: {real_distance:.1f}mm >= {LASER_STOP_OPEN}mm)")
                                    print(f"      当前位置: {current_actual_pos:.6f}")
                                    print(f"      等待位置稳定 (变化<{POS_STABLE_THRESHOLD})...")
                            else:  # 闭合
                                if real_distance <= LASER_STOP_CLOSE:
                                    target_reached = True
                                    target_reached_time = time.time()
                                    last_stable_pos = current_actual_pos
                                    print(f"\n    ✓ 到达闭合目标位置 (激光: {real_distance:.1f}mm <= {LASER_STOP_CLOSE}mm)")
                                    print(f"      当前位置: {current_actual_pos:.6f}")
                                    print(f"      等待位置稳定 (变化<{POS_STABLE_THRESHOLD})...")
                    else:
                        # 基于堵转检测的停止（阶段2使用）
                        pos_change = abs(current_actual_pos - last_actual_pos)
                        if pos_change < pos_threshold:
                            stall_count += 1
                            if stall_count >= stable_target:
                                target_reached = True
                                target_reached_time = time.time()
                                last_stable_pos = current_actual_pos
                                print(f"\n    ✓ 检测到堵转，到达机械限位")
                                print(f"      当前位置: {current_actual_pos:.6f}")
                                print(f"      连续稳定次数: {stall_count}/{stable_target}")
                                print(f"      等待位置稳定 (变化<{POS_STABLE_THRESHOLD})...")
                        else:
                            if stall_count > 0:
                                print(f"    ↗ 位置变化恢复: {pos_change:.8f}, 重置堵转计数")
                            stall_count = 0
                
                # 6. 实时显示（每0.5秒）
                current_time = time.time()
                if current_time - last_print_time >= 0.5:
                    elapsed = current_time - start_time
                    laser_str = f", 激光: {real_distance:.1f}mm" if real_distance is not None else ""
                    step_str = f", 步长: {current_step:.6f}" if slowdown_active else ""
                    if target_reached:
                        stable_progress = f"{pos_stable_count}/{POS_STABLE_COUNT}"
                        status_str = f", 稳定计数: {stable_progress}"
                    elif use_laser_stop:
                        status_str = f", 运动中"
                    else:
                        status_str = f", 堵转计数: {stall_count}/{stable_target}"
                    print(f"    [{elapsed:5.1f}s] 位置: {current_actual_pos:8.6f}, "
                          f"指令: {command_pos:8.6f}{laser_str}{step_str}{status_str}")
                    last_print_time = current_time
                
                # 7. 检查位置是否真正稳定或超时
                if target_reached and target_reached_time is not None:
                    # 检查位置是否稳定
                    if pos_stable_count >= POS_STABLE_COUNT:
                        elapsed = time.time() - start_time
                        print(f"\n    ✓ 标定完成！位置已稳定")
                        print(f"      最终位置: {current_actual_pos:.6f}")
                        print(f"      最终激光距离: {real_distance:.1f}mm" if real_distance is not None else "")
                        print(f"      总耗时: {elapsed:.2f}秒")
                        return current_actual_pos
                    
                    # 检查是否超时
                    if time.time() - target_reached_time >= MAX_WAIT_TIME:
                        elapsed = time.time() - start_time
                        print(f"\n    ! 标定完成（等待超时）")
                        print(f"      最终位置: {current_actual_pos:.6f}")
                        print(f"      最终激光距离: {real_distance:.1f}mm" if real_distance is not None else "")
                        print(f"      总耗时: {elapsed:.2f}秒")
                        return current_actual_pos
                
                last_actual_pos = current_actual_pos
                
        except KeyboardInterrupt:
            print("\n    ! 用户中断标定")
            # 停止运动
            current_pos = self._get_feedback_scalar(pos_name)
            if current_pos is not None:
                self.g.robot.set_actions({part: {"type": "position", "position": [current_pos]}})
            raise

    def _calibrate_gripper_climb_to_laser(self, part: str, pos_name: str, direction_sign: int,
                                           target_laser: float, laser_tolerance: float = 0.2,
                                           target_speed: float = 0.3, ctrl_freq: int = 100,
                                           use_jam_detection: bool = False,
                                           jam_threshold: float = 0.0001,
                                           jam_count: int = 50) -> float:
        """
        通过位置爬坡移动到指定的激光距离目标（阶段2专用）
        支持两种停止模式：激光检测或堵转检测
        
        Args:
            part: 夹爪部件名
            pos_name: 位置反馈数据名
            direction_sign: -1 (向内闭合), 1 (向外打开)
            target_laser: 目标激光距离（mm）
            laser_tolerance: 激光距离容差（mm），默认±0.2mm
            target_speed: 目标速度 (单位/秒)，默认0.3
            ctrl_freq: 控制频率 (Hz)
            use_jam_detection: True=使用堵转检测停止, False=使用激光检测停止
            jam_threshold: 堵转检测阈值（位置变化小于此值认为堵转）
            jam_count: 连续多少次位置不变认为堵转
            
        Returns:
            最终位置值
        """
        dt = 1.0 / ctrl_freq
        base_step = target_speed * dt
        current_step = base_step
        
        direction_str = "闭合" if direction_sign == -1 else "张开"
        stop_mode_str = "堵转检测" if use_jam_detection else "激光检测"
        print(f"\n>>> 开始移动{direction_str}到目标 (停止模式: {stop_mode_str})")
        if not use_jam_detection:
            print(f"    激光目标: {target_laser}mm (容差: ±{laser_tolerance}mm)")
        print(f"    控制频率: {ctrl_freq}Hz, 目标速度: {target_speed}")
        print(f"    基础步长: {base_step:.6f}")
        
        # 获取起始位置
        start_pos = self._get_feedback_scalar(pos_name)
        if start_pos is None:
            start_pos = 0.0
        command_pos = start_pos
        last_actual_pos = start_pos
        
        # 减速阈值（接近目标时减速）- 仅激光模式使用
        SLOWDOWN_DISTANCE = 2.0  # 距离目标2mm时开始减速
        STEP_REDUCTION_FACTOR = 0.1  # 减速因子
        
        # 位置稳定检测参数
        POS_STABLE_THRESHOLD = 0.0001  # 位置变化小于此值认为稳定
        POS_STABLE_COUNT = 50          # 连续50次（0.5秒）稳定认为真正停止
        MAX_WAIT_TIME = 5.0            # 最大等待时间5秒
        
        # 堵转检测参数
        jam_stable_count = 0  # 堵转计数器
        
        print(f"    起始位置: {start_pos:.6f}")
        print("    开始位置爬坡... (按Ctrl+C可中断)")
        
        last_print_time = time.time()
        start_time = time.time()
        slowdown_active = False
        target_reached = False
        target_reached_time = None
        pos_stable_count = 0
        last_stable_pos = None
        
        try:
            while True:
                # 如果已经到达目标位置，不再下发任何指令，只等待稳定
                if target_reached:
                    # 只读取位置和激光数据，用于显示
                    current_actual_pos = self._get_feedback_scalar(pos_name)
                    if current_actual_pos is None:
                        current_actual_pos = last_actual_pos
                    real_distance = self._get_feedback_scalar("real_distance")
                    
                    # 持续发送当前位置作为目标，确保电机停止运动
                    self.g.robot.set_actions({part: {"type": "position", "position": [current_actual_pos]}})
                    
                    # 检查位置是否稳定
                    if last_stable_pos is not None:
                        pos_change = abs(current_actual_pos - last_stable_pos)
                        if pos_change < POS_STABLE_THRESHOLD:
                            pos_stable_count += 1
                        else:
                            pos_stable_count = 0
                    last_stable_pos = current_actual_pos
                    
                    time.sleep(dt)
                else:
                    # 获取激光测距仪数据（用于显示和减速）
                    real_distance = self._get_feedback_scalar("real_distance")
                    
                    # 根据距离目标的远近调整步长（仅激光模式）
                    if not use_jam_detection and real_distance is not None:
                        distance_to_target = abs(real_distance - target_laser)
                        if distance_to_target < SLOWDOWN_DISTANCE:  # 接近目标
                            if not slowdown_active:
                                current_step = base_step * STEP_REDUCTION_FACTOR
                                slowdown_active = True
                                print(f"    ↓ 接近目标，步长减小至 {current_step:.6f} (距离: {distance_to_target:.2f}mm)")
                        else:
                            if slowdown_active:
                                current_step = base_step
                                slowdown_active = False
                                print(f"    ↑ 恢复正常步长 {current_step:.6f}")
                    
                    # 1. 更新指令位置
                    command_pos += current_step * direction_sign
                    
                    # 2. 发送位置指令
                    self.g.robot.set_actions({part: {"type": "position", "position": [command_pos]}})
                    
                    # 3. 等待一个周期
                    time.sleep(dt)
                    
                    # 4. 获取当前实际位置
                    current_actual_pos = self._get_feedback_scalar(pos_name)
                    if current_actual_pos is None:
                        current_actual_pos = last_actual_pos
                    
                    # 5. 停止检测
                    if use_jam_detection:
                        # 基于堵转的停止检测
                        pos_change = abs(current_actual_pos - last_actual_pos)
                        if pos_change < jam_threshold:
                            jam_stable_count += 1
                            if jam_stable_count >= jam_count:
                                target_reached = True
                                target_reached_time = time.time()
                                last_stable_pos = current_actual_pos
                                print(f"\n    ✓ 检测到堵转 (连续{jam_count}次位置变化<{jam_threshold})")
                                print(f"      当前位置: {current_actual_pos:.6f}")
                                if real_distance is not None:
                                    print(f"      当前激光: {real_distance:.2f}mm")
                                print(f"      等待位置稳定...")
                        else:
                            jam_stable_count = 0
                    else:
                        # 基于激光距离的停止检测
                        if real_distance is not None:
                            if abs(real_distance - target_laser) <= laser_tolerance:
                                target_reached = True
                                target_reached_time = time.time()
                                last_stable_pos = current_actual_pos
                                print(f"\n    ✓ 到达激光目标位置 (激光: {real_distance:.2f}mm, 目标: {target_laser}mm ±{laser_tolerance}mm)")
                                print(f"      当前位置: {current_actual_pos:.6f}")
                                print(f"      等待位置稳定...")
                
                # 6. 实时显示（每0.5秒）
                current_time = time.time()
                if current_time - last_print_time >= 0.5:
                    elapsed = current_time - start_time
                    laser_str = f", 激光: {real_distance:.2f}mm" if real_distance is not None else ""
                    step_str = f", 步长: {current_step:.6f}" if slowdown_active else ""
                    if use_jam_detection and not target_reached:
                        jam_str = f", 堵转计数: {jam_stable_count}/{jam_count}"
                    else:
                        jam_str = ""
                    if target_reached:
                        stable_progress = f"{pos_stable_count}/{POS_STABLE_COUNT}"
                        status_str = f", 稳定计数: {stable_progress}"
                    else:
                        status_str = f", 运动中"
                    print(f"    [{elapsed:5.1f}s] 位置: {current_actual_pos:8.6f}, "
                          f"指令: {command_pos:8.6f}{laser_str}{step_str}{jam_str}{status_str}")
                    last_print_time = current_time
                
                # 7. 检查位置是否真正稳定或超时
                if target_reached and target_reached_time is not None:
                    # 检查位置是否稳定
                    if pos_stable_count >= POS_STABLE_COUNT:
                        elapsed = time.time() - start_time
                        print(f"\n    ✓ 移动完成！位置已稳定")
                        print(f"      最终位置: {current_actual_pos:.6f}")
                        print(f"      最终激光距离: {real_distance:.2f}mm" if real_distance is not None else "")
                        print(f"      总耗时: {elapsed:.2f}秒")
                        return current_actual_pos
                    
                    # 检查是否超时
                    if time.time() - target_reached_time >= MAX_WAIT_TIME:
                        elapsed = time.time() - start_time
                        print(f"\n    ! 移动完成（等待超时）")
                        print(f"      最终位置: {current_actual_pos:.6f}")
                        print(f"      最终激光距离: {real_distance:.2f}mm" if real_distance is not None else "")
                        print(f"      总耗时: {elapsed:.2f}秒")
                        return current_actual_pos
                
                last_actual_pos = current_actual_pos
                
        except KeyboardInterrupt:
            print("\n    ! 用户中断移动")
            # 停止运动
            current_pos = self._get_feedback_scalar(pos_name)
            if current_pos is not None:
                self.g.robot.set_actions({part: {"type": "position", "position": [current_pos]}})
            raise

    def _smooth_move_to(self, part: str, pos_name: str, target: float, duration: float = 2.0, 
                         use_laser_check: bool = False, laser_target: float = None, laser_tolerance: float = 1.0):
        """
        平滑地将夹爪移动到目标位置
        
        Args:
            part: 夹爪部件名
            pos_name: 位置反馈数据名
            target: 目标位置
            duration: 移动持续时间（秒）
            use_laser_check: 是否使用激光距离作为到位判断
            laser_target: 激光目标距离（mm）
            laser_tolerance: 激光距离容差（mm）
        """
        steps = int(duration * 100)  # 100Hz
        dt = 0.01
        
        start_pos = self._get_feedback_scalar(pos_name)
        if start_pos is None:
            start_pos = 0.0
        
        print(f"\n>>> 平滑移动至目标: {target:.6f} (预计耗时 {duration}s)")
        print(f"    起始位置: {start_pos:.6f}")
        
        # 第一阶段：发送平滑插值指令
        for i in range(steps + 1):
            alpha = i / float(steps)
            curr_cmd = start_pos + (target - start_pos) * alpha
            self.g.robot.set_actions({part: {"type": "position", "position": [curr_cmd]}})
            time.sleep(dt)
        
        # 第二阶段：等待实际到达目标位置
        print("    等待到位...")
        wait_start = time.time()
        max_wait = 5.0  # 增加最大等待时间到5秒
        
        while time.time() - wait_start < max_wait:
            current_pos = self._get_feedback_scalar(pos_name)
            
            # 如果使用激光检查
            if use_laser_check and laser_target is not None:
                real_distance = self._get_feedback_scalar("real_distance")
                if real_distance is not None and abs(real_distance - laser_target) <= laser_tolerance:
                    print(f"    ✓ 激光到位: {real_distance:.1f}mm (目标: {laser_target}mm)")
                    return
            
            # 位置检查
            if current_pos is not None and abs(current_pos - target) <= 0.001:
                print(f"    ✓ 移动到位: {current_pos:.6f}")
                return
                
            self.g.robot.set_actions({part: {"type": "position", "position": [target]}})
            time.sleep(dt)
        
        # 超时后的最终状态
        final_pos = self._get_feedback_scalar(pos_name)
        final_laser = self._get_feedback_scalar("real_distance")
        print(f"    ! 等待到位超时")
        print(f"      最终位置: {final_pos:.6f}" if final_pos is not None else "      最终位置: N/A")
        if final_laser is not None:
            print(f"      最终激光: {final_laser:.1f}mm")

    @hide_ui_while
    def full_auto_calibration(self, part: str, pos_name: str = "gripper_pos"):
        """
        【参考已有完整流程】自动标定并设零（分阶段版）
        
        阶段1 - 标定并计算参数（步骤0-4）：
        1. 自动重置yaml参数为 length_per_radian=1.0, offset_at_hardware_zero=0.0
        2. 寻找正向极限（全开）→ 停2秒 → 读取 laser_open
        3. 寻找负向极限（闭合）→ 停2秒 → 读取 laser_close
        4. 计算 length_per_radian 并写入yaml，同步到UI，保存状态后退出
        
        阶段2 - 完成设零（步骤5-8，重新进入后执行）：
        5. 在闭合位置设零
        6. 移动到 0.025 位置
        7. 再次设零
        8. 设置 offset_at_hardware_zero = 0.025
        """
        # 检查是否有保存的状态（阶段2）
        if hasattr(self, '_calibration_state') and self._calibration_state.get('part') == part:
            return self._full_auto_calibration_phase2(part, pos_name)
        
        # ========== 阶段1: 标定并计算参数 ==========
        print("=" * 60)
        print("【参考已有完整流程】自动标定并设零（阶段1: 标定并计算参数）")
        print("=" * 60)
        print(f"夹爪: {part}")
        print()
        
        # 步骤0: 重置参数
        print(">>> 步骤0: 重置yaml参数为默认值")
        try:
            yaml_path = self._write_gripper_yaml_params(
                part=part,
                length_per_radian=1.0,
                offset_at_hardware_zero=0.0,
            )
            print(f"    ✓ 参数已重置: length_per_radian=1.0, offset_at_hardware_zero=0.0")
            print(f"      文件: {yaml_path}")
        except Exception as e:
            print(f"    ✗ 重置yaml失败: {e}")
            self.g.loggerUI.error(f"重置yaml失败: {e}")
            input("按回车返回")
            return
        
        # 重新加载
        if hasattr(self.g, "reload_robot_from_yaml"):
            try:
                self.g.reload_robot_from_yaml()
                print(f"    ✓ 重新加载完成")
            except Exception as e:
                print(f"    ✗ 重新加载失败: {e}")
                input("按回车返回")
                return
        else:
            print("    ! reload_robot_from_yaml() 未找到，请手动重启")
            input("按回车返回")
            return
        
        # 等待硬件就绪
        print("\n>>> 等待硬件启动...")
        max_wait = 30
        wait_start = time.time()
        while time.time() - wait_start < max_wait:
            test_pos = self._get_feedback_scalar(pos_name)
            if test_pos is not None:
                print(f"    ✓ 硬件已就绪")
                break
            time.sleep(0.5)
        else:
            print("    ✗ 等待硬件就绪超时")
            input("按回车返回")
            return
        
        # 取消限位（允许超限运动）
        print(">>> 取消硬件限位...")
        try:
            self.g.robot.send_command(part, {"command": "set_limit", "enabled": [False], "lower": [-0.5], "upper": [0.5]})
            print(f"    ✓ 限位已取消")
        except Exception as e:
            print(f"    ! 取消限位失败（可能不影响）: {e}")

        # ========== 准备步骤: 闭合到真正闭合位置 + 激光归零 ==========
        print("\n" + "=" * 60)
        print(">>> 准备步骤: 闭合到真正闭合位置（堵转检测）+ 激光归零")
        print("=" * 60)
        try:
            true_close_pos = self._calibrate_gripper_climb_to_laser(
                part=part,
                pos_name=pos_name,
                direction_sign=-1,
                target_laser=0.0,
                laser_tolerance=0.2,
                target_speed=0.3,
                ctrl_freq=100,
                use_jam_detection=True,
                jam_threshold=0.0001,
                jam_count=50,
            )
            print(f"    ✓ 已到达真正闭合位置: {true_close_pos:.6f}")
            self.g.loggerUI.info(f"[准备步骤] {part}: true_close_pos={true_close_pos:.6f}")
        except Exception as e:
            print(f"    ✗ 闭合到真正闭合位置失败: {e}")
            self.g.loggerUI.error(f"[准备步骤] 闭合失败: {e}")
            input("按回车返回")
            return

        try:
            self._set_lasers_zero_core()
            print("    ✓ 激光归零完成")
            self.g.loggerUI.info(f"[准备步骤] {part}: laser set_zero done")
        except Exception as e:
            print(f"    ✗ 激光归零失败: {e}")
            self.g.loggerUI.error(f"[准备步骤] 激光归零失败: {e}")
            input("按回车返回")
            return

        # ========== 步骤1: 寻找正向极限（全开） ==========
        print("\n" + "=" * 60)
        print(">>> 步骤1: 寻找正向极限（全开）")
        print("=" * 60)
        
        try:
            max_pos = self._calibrate_gripper_climb(part, pos_name, direction_sign=1)
        except KeyboardInterrupt:
            print("\n标定被用户中断")
            input("按回车返回")
            return
        except Exception as e:
            print(f"\n标定失败: {e}")
            self.g.loggerUI.error(f"张开标定失败: {e}")
            input("按回车返回")
            return
        
        print(f"\n    张开位置记录: max_pos = {max_pos:.6f}")
        
        # 停2秒，读取激光数据
        print("    稳定2秒后读取激光数据...")
        time.sleep(2.0)
        laser_open = self._get_feedback_scalar("real_distance")
        print(f"    激光测距（张开）: laser_open = {laser_open:.6f} mm")
        
        # ========== 步骤2: 寻找负向极限（闭合） ==========
        print("\n" + "=" * 60)
        print(">>> 步骤2: 寻找负向极限（闭合）")
        print("=" * 60)
        
        try:
            min_pos = self._calibrate_gripper_climb(part, pos_name, direction_sign=-1)
        except KeyboardInterrupt:
            print("\n标定被用户中断")
            input("按回车返回")
            return
        except Exception as e:
            print(f"\n标定失败: {e}")
            self.g.loggerUI.error(f"闭合标定失败: {e}")
            input("按回车返回")
            return
        
        print(f"\n    闭合位置记录: min_pos = {min_pos:.6f}")
        
        # 停2秒，读取激光数据
        print("    稳定2秒后读取激光数据...")
        time.sleep(2.0)
        laser_close = self._get_feedback_scalar("real_distance")
        print(f"    激光测距（闭合）: laser_close = {laser_close:.6f} mm")
        
        # ========== 步骤3: 显示标定报告 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤3: 标定报告")
        print("=" * 60)
        
        stroke_rad = abs(max_pos - min_pos)
        real_distance = laser_open - laser_close if laser_open and laser_close else 0.0
        
        print(f"\n    标定结果:")
        print(f"    ┌─────────────────────────────────────┐")
        print(f"    │  张开位置 (max): {max_pos:12.6f}     │")
        print(f"    │  激光张开:       {laser_open:12.6f} mm  │") if laser_open else print(f"    │  激光张开:       未读取               │")
        print(f"    │  闭合位置 (min): {min_pos:12.6f}     │")
        print(f"    │  激光闭合:       {laser_close:12.6f} mm  │") if laser_close else print(f"    │  激光闭合:       未读取               │")
        print(f"    │  弧度行程 (rad): {stroke_rad:11.6f}     │")
        print(f"    │  实际行程 (mm):  {real_distance:11.6f}     │") if real_distance else print(f"    │  实际行程 (mm):  未计算               │")
        print(f"    └─────────────────────────────────────┘")
        
        self.g.loggerUI.info(f"[标定完成] {part}: max={max_pos:.6f}, min={min_pos:.6f}, stroke_rad={stroke_rad:.6f}")
        self.g.loggerUI.info(
            f"[标定数据] {part}: laser_open={laser_open}, laser_close={laser_close}, "
            f"stroke_rad={stroke_rad:.6f}, real_distance={real_distance}"
        )
        
        # 检查激光数据有效性
        if laser_open is None or laser_close is None:
            print("\n    ✗ 错误: 未能读取激光测距数据")
            input("按回车返回")
            return
        
        if real_distance <= 0:
            print(f"\n    ✗ 错误: 实际行程距离无效 ({real_distance:.6f})")
            input("按回车返回")
            return
        
        if stroke_rad < 1e-9:
            print(f"\n    ✗ 错误: 弧度变化过小 ({stroke_rad:.10f})")
            input("按回车返回")
            return
        
        # ========== 步骤4: 计算并写入参数 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤4: 计算并写入 length_per_radian")
        print("=" * 60)
        
        try:
            # 计算 length_per_radian
            length_per_radian = real_distance / (2000.0 * stroke_rad)
            length_per_radian_10 = round(length_per_radian, 10)
            
            print(f"\n计算结果:")
            print(f"  实际行程: {real_distance:.6f} mm")
            print(f"  弧度变化: {stroke_rad:.6f} rad")
            print(f"  length_per_radian = {length_per_radian_10:.10f}")
            self.g.loggerUI.info(
                f"[参数计算] {part}: laser_open={laser_open}, laser_close={laser_close}, "
                f"stroke_rad={stroke_rad:.6f}, real_distance={real_distance:.6f}, "
                f"length_per_radian={length_per_radian_10:.10f}"
            )
            
            # 写入yaml
            yaml_path = self._write_gripper_yaml_params(
                part=part,
                length_per_radian=length_per_radian_10,
            )
            print(f"\n    ✓ length_per_radian 已写入: {length_per_radian_10:.10f}")
            print(f"      文件: {yaml_path}")
            
            # 同步到UI
            if hasattr(self.g, "reload_robot_from_yaml"):
                self.g.reload_robot_from_yaml()
                print(f"    ✓ 参数已同步到UI")
            else:
                print(f"    ! reload_robot_from_yaml() 未找到，请手动重启")
            
            self.g.loggerUI.info(f"[参数更新] {part}: length_per_radian={length_per_radian_10:.10f}")
            
        except Exception as e:
            print(f"\n    ✗ 计算或写入失败: {e}")
            self.g.loggerUI.error(f"计算或写入失败: {e}")
            input("按回车返回")
            return
        
        # 保存状态，直接进入阶段2
        self._calibration_state = {
            'part': part,
            'min_pos': min_pos,
            'max_pos': max_pos,
            'laser_open': laser_open,
            'laser_close': laser_close,
            'length_per_radian': length_per_radian_10,
            'laser_close_target': 5.0,  # 闭合时的激光目标距离
        }
        
        # ========== 阶段1完成，直接进入阶段2 ==========
        print("\n" + "=" * 60)
        print("【阶段1完成】参数已计算并写入，即将进入阶段2")
        print("=" * 60)
        print(f"\n    标定数据:")
        print(f"      length_per_radian = {length_per_radian_10:.10f}")
        print(f"      min_pos = {min_pos:.6f}")
        print(f"      max_pos = {max_pos:.6f}")
        
        # 直接进入阶段2
        return self._full_auto_calibration_phase2(part, pos_name)


    def _full_auto_calibration_phase2(self, part: str, pos_name: str = "gripper_pos"):
        """
        阶段2: 完成设零流程
        """
        state = self._calibration_state
        min_pos = state['min_pos']
        length_per_radian_10 = state['length_per_radian']
        
        print("\n" + "=" * 60)
        print("【参考已有完整流程】自动标定并设零（阶段2: 完成设零）")
        print("=" * 60)
        print(f"夹爪: {part}")
        print(f"\n    恢复的标定数据:")
        print(f"      length_per_radian = {length_per_radian_10:.10f}")
        print(f"      min_pos = {min_pos:.6f} (阶段1记录值，仅供参考)")
        
        # ========== 步骤4.5: 等待系统稳定并同步当前位置 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤4.5: 等待系统稳定并同步当前位置")
        print("=" * 60)
        print("\n    重新加载参数后，等待电机位置反馈稳定...")
        time.sleep(0.5)  # 等待0.5秒让系统稳定
        
        # 读取当前实际位置（改完length_per_radian后的真实位置）
        current_actual_pos = self._get_feedback_scalar(pos_name)
        current_laser = self._get_feedback_scalar("real_distance")
        print(f"\n    读取到的位置: {current_actual_pos:.6f}")
        if current_laser is not None:
            print(f"    当前激光距离: {current_laser:.2f}mm")
        
        # 检查位置是否合理（如果接近min_pos，说明需要同步）
        if abs(current_actual_pos - min_pos) < 0.01:
            print(f"\n    ! 警告: 位置读数({current_actual_pos:.6f})接近阶段1的min_pos({min_pos:.6f})")
            print(f"      这可能是因为电机控制器缓存了旧的位置值")
            print(f"      尝试通过发送当前位置指令来同步...")
            
            # 发送一个小的位置指令来刷新电机控制器的状态
            # 先读取激光距离，确认实际位置
            if current_laser is not None:
                # 根据激光距离估算应该的位置值
                # 激光5mm对应的位置大约是 min_pos * length_per_radian_10 ≈ 0.0015
                expected_pos = min_pos * length_per_radian_10
                print(f"\n    根据激光距离估算，当前位置应该是: {expected_pos:.6f}")
                print(f"    发送位置指令来同步电机控制器...")
                
                # 发送当前估算位置作为目标
                self.g.robot.set_actions({part: {"type": "position", "position": [expected_pos]}})
                time.sleep(0.5)  # 等待指令生效
                
                # 再次读取位置
                new_pos = self._get_feedback_scalar(pos_name)
                print(f"    同步后位置: {new_pos:.6f}")
            
        print(f"\n    注意: 由于length_per_radian改变，位置数值已变化")
        print(f"          阶段1的min_pos({min_pos:.6f})已不再适用")
        
        # ========== 步骤5: 移动到真正闭合位置（使用堵转检测） ==========
        print("\n" + "=" * 60)
        print(">>> 步骤5: 移动到真正闭合位置（使用堵转检测）")
        print("=" * 60)
        
        # 从当前实际位置开始，闭合直到堵转（机械闭合）
        print(f"\n    从当前位置继续闭合直到机械堵转...")
        try:
            # 使用爬坡方式闭合，通过堵转检测判断是否到达真正闭合位置
            true_close_pos = self._calibrate_gripper_climb_to_laser(
                part=part,
                pos_name=pos_name,
                direction_sign=-1,  # 闭合方向
                target_laser=0.0,   # 激光目标（仅用于显示）
                laser_tolerance=0.2, # 激光容差（仅用于显示）
                target_speed=0.3,   # 目标速度
                ctrl_freq=100,
                use_jam_detection=True,  # 使用堵转检测
                jam_threshold=0.0001,    # 堵转阈值
                jam_count=50,            # 连续50次（0.5秒）位置不变认为堵转
            )
            print(f"    ✓ 已到达真正闭合位置: {true_close_pos:.6f}")
            self.g.loggerUI.info(
                f"[阶段2-步骤5] {part}: true_close_pos={true_close_pos:.6f}"
            )
        except Exception as e:
            print(f"    ✗ 移动到真正闭合位置失败: {e}")
            self.g.loggerUI.error(f"移动到真正闭合位置失败: {e}")
            input("按回车返回")
            return
        
        # ========== 步骤6: 在真正闭合位置设零 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤6: 在真正闭合位置设零")
        print("=" * 60)
        
        print(f"\n    执行硬件设零...")
        try:
            pos_before_zero = self._get_feedback_scalar(pos_name)
            self.set_zero(part=part)
            pos_after_zero = self._get_feedback_scalar(pos_name)
            print(f"    ✓ 设零完成！")
            self.g.loggerUI.info(f"[设零完成] {part} 在真正闭合位置设零")
            self.g.loggerUI.info(
                f"[阶段2-步骤6] {part}: pos_before_zero={pos_before_zero}, pos_after_zero={pos_after_zero}"
            )
        except Exception as e:
            print(f"    ✗ 设零失败: {e}")
            self.g.loggerUI.error(f"设零失败: {e}")
            input("按回车返回")
            return
        
        # ========== 步骤7: 移动到 0.025 位置 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤7: 移动到 0.025 位置")
        print("=" * 60)
        
        target_pos = 0.025
        print(f"\n    目标位置: {target_pos:.6f}")
        print(f"    使用平滑移动...")
        
        try:
            # 使用 _smooth_move_to 进行平滑移动
            self._smooth_move_to(part, pos_name, target_pos, duration=2.0)
            time.sleep(0.5)
            current_pos = self._get_feedback_scalar(pos_name)
            current_laser = self._get_feedback_scalar("real_distance")
            if current_pos is None:
                print(f"    ! 未读取到当前位置，回退使用目标值 0.025")
                current_pos = target_pos
            print(f"    ✓ 已到达位置: {current_pos:.6f}")
            if current_laser is not None:
                print(f"    当前激光距离: {current_laser:.2f}mm")
            self.g.loggerUI.info(
                f"[阶段2-步骤7] {part}: target_pos={target_pos:.6f}, "
                f"actual_pos={current_pos:.6f}, laser={current_laser}"
            )

            # 闭环微调：依据激光距离收敛到 50±0.2mm
            TARGET_LASER = 50.0
            LASER_TOL = 0.2
            MAX_ITERS = 20
            STEP = 0.0002  # 小步进，避免过冲
            if current_laser is None:
                print("    ! 无法读取激光距离，跳过闭环微调")
            else:
                for i in range(1, MAX_ITERS + 1):
                    if abs(current_laser - TARGET_LASER) <= LASER_TOL:
                        print(f"    ✓ 激光到位: {current_laser:.2f}mm (迭代{ i })")
                        break
                    if current_laser > TARGET_LASER + LASER_TOL:
                        # 太大，向闭合方向微调
                        current_pos -= STEP
                    else:
                        # 太小，向张开方向微调
                        current_pos += STEP
                    self._hold_position_command(part=part, target=current_pos, hold_s=0.2)
                    time.sleep(0.05)
                    current_laser = self._get_feedback_scalar("real_distance")
                    print(f"    [微调{i:02d}] pos={current_pos:.6f}, laser={current_laser:.2f}mm")
                    self.g.loggerUI.info(
                        f"[阶段2-步骤7-微调{i:02d}] {part}: pos={current_pos:.6f}, laser={current_laser}"
                    )
                else:
                    print(f"    ! 微调结束仍未到位: laser={current_laser:.2f}mm")
            # 保留微调后的位置作为 offset 写入值
        except Exception as e:
            print(f"    ✗ 移动失败: {e}")
            self.g.loggerUI.error(f"移动失败: {e}")
            input("按回车返回")
            return
        
        # ========== 步骤8: 再次设零 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤8: 再次设零（在 0.025 位置）")
        print("=" * 60)
        
        try:
            print(f"\n    执行硬件设零...")
            pos_before_zero = self._get_feedback_scalar(pos_name)
            self.set_zero(part=part)
            pos_after_zero = self._get_feedback_scalar(pos_name)
            print(f"    ✓ 设零完成！")
            self.g.loggerUI.info(f"[设零完成] {part} 在 0.025 位置设零")
            self.g.loggerUI.info(
                f"[阶段2-步骤8] {part}: pos_before_zero={pos_before_zero}, pos_after_zero={pos_after_zero}"
            )
        except Exception as e:
            print(f"    ✗ 设零失败: {e}")
            self.g.loggerUI.error(f"设零失败: {e}")
            input("按回车返回")
            return
        
        # ========== 步骤9: 设置 offset_at_hardware_zero = 0.025 ==========
        print("\n" + "=" * 60)
        print(">>> 步骤9: 设置 offset_at_hardware_zero = 0.025")
        print("=" * 60)
        
        try:
            current_pos = 0.025
            yaml_path = self._write_gripper_yaml_params(
                part=part,
                offset_at_hardware_zero=current_pos,
            )
            print(f"    ✓ offset_at_hardware_zero 已设置为 {current_pos:.6f}")
            print(f"      文件: {yaml_path}")
            self.g.loggerUI.info(
                f"[阶段2-步骤9] {part}: offset_at_hardware_zero={current_pos:.6f}"
            )
            
            # 同步到UI
            if hasattr(self.g, "reload_robot_from_yaml"):
                self.g.reload_robot_from_yaml()
                print(f"    ✓ 参数已同步到UI")
            else:
                print(f"    ! reload_robot_from_yaml() 未找到，请手动重启")
            
            self.g.loggerUI.info(f"[参数更新] {part}: offset_at_hardware_zero={current_pos:.6f}")
        except Exception as e:
            print(f"    ✗ 写入或同步失败: {e}")
            self.g.loggerUI.error(f"写入或同步失败: {e}")
        
        # ========== 步骤10: 验证最终位置（激光距离应为50±0.2mm） ==========
        print("\n" + "=" * 60)
        print(">>> 步骤10: 验证最终位置")
        print("=" * 60)
        
        # 等待一段时间让系统稳定
        print("\n    等待系统稳定...")
        time.sleep(1.0)
        
        # 读取当前激光距离
        final_laser = self._get_feedback_scalar("real_distance")
        if final_laser is not None:
            TARGET_LASER = 50.0  # 目标激光距离50mm
            LASER_TOLERANCE = 0.2  # 容差±0.2mm
            
            print(f"\n    当前激光距离: {final_laser:.2f}mm")
            print(f"    目标范围: {TARGET_LASER}±{LASER_TOLERANCE}mm ({TARGET_LASER - LASER_TOLERANCE} ~ {TARGET_LASER + LASER_TOLERANCE}mm)")
            
            if abs(final_laser - TARGET_LASER) <= LASER_TOLERANCE:
                print(f"    ✓ 验证通过！激光距离在合格范围内")
                self.g.loggerUI.info(f"[验证通过] {part} 最终激光距离: {final_laser:.2f}mm (目标: {TARGET_LASER}±{LASER_TOLERANCE}mm)")
            else:
                print(f"    ✗ 验证失败！激光距离超出合格范围")
                print(f"      偏差: {final_laser - TARGET_LASER:.2f}mm")
                self.g.loggerUI.error(f"[验证失败] {part} 最终激光距离: {final_laser:.2f}mm，超出目标范围 {TARGET_LASER}±{LASER_TOLERANCE}mm")
                print(f"\n    ! 警告: 标定结果可能不合格，请检查机械结构或重新标定")
            self.g.loggerUI.info(
                f"[阶段2-步骤10] {part}: final_laser={final_laser:.2f}mm, "
                f"target={TARGET_LASER:.2f}mm, diff={final_laser - TARGET_LASER:.2f}mm"
            )
        else:
            print(f"    ! 无法读取激光距离，跳过验证")
            self.g.loggerUI.warn(f"[验证跳过] {part} 无法读取激光距离")

        try:
            self.g.robot.send_command(
                part,
                {"command": "set_limit", "enabled": [True], "lower": [0.0], "upper": [0.05]},
            )
            print("    ✓ 已设置软限位: 0.0 ~ 0.05")
            self.g.loggerUI.info(f"[限位设置] {part}: 0.0~0.05")
        except Exception as e:
            print(f"    ! 设置软限位失败: {e}")
            self.g.loggerUI.warn(f"[限位设置失败] {part}: {e}")
        
        # 清除状态
        delattr(self, '_calibration_state')
        
        # ========== 完成 ==========
        print("\n" + "=" * 60)
        print("【参考已有完整流程】自动标定并设零 - 全部完成")
        print("=" * 60)
        print(f"\n    最终参数:")
        print(f"      length_per_radian = {length_per_radian_10:.10f}")
        print(f"      offset_at_hardware_zero = 0.025")
        print(f"      零点位置 = 0.025 位置（从闭合位置张开）")
        
        input("\n按回车返回")


    @hide_ui_while
    def manual_calibration_open_to_max(self, part: str, pos_name: str = "gripper_pos"):
        upper = None
        try:
            limit_feed = self.g.robot.send_command(part, {"command": "get_limit"})
            upper = self._coerce_scalar(limit_feed.get("upper") if isinstance(limit_feed, dict) else None)
        except Exception as e:
            self.g.loggerUI.warn(f"get_limit failed: {e}")

        default_target = upper if upper is not None else 1.0
        raw_target = input(f"open target (default {default_target:.6f}): ").strip()
        target = float(raw_target) if raw_target else float(default_target)

        raw_hold = input("hold seconds at open target (default 3.0): ").strip()
        hold_s = float(raw_hold) if raw_hold else 3.0

        try:
            self.g.robot.send_command(part, {"command": "set_control_mode", "mode": "position"})
        except Exception:
            pass

        print(f"open gripper to {target:.6f}, hold {hold_s:.2f}s")
        self._hold_position_command(part=part, target=target, hold_s=hold_s)

        cur = self._get_feedback_scalar(pos_name)
        self.g.loggerUI.info(f"manual open done: target={target:.6f}, current={cur}")
        print(f"manual open done, current pos={cur}")
        input("Press Enter to return")

    @hide_ui_while
    def recover_mobility(self, part: str, pos_name: str = "gripper_pos"):
        print("Recovery: try to restore gripper mobility")
        try:
            self.g.robot.send_command(part, {"command": "set_control_mode", "mode": "position"})
        except Exception as e:
            self.g.loggerUI.warn(f"set_control_mode(position) failed: {e}")

        cur = self._get_feedback_scalar(pos_name)
        if cur is None:
            cur = 0.0

        self._hold_position_command(part=part, target=cur, hold_s=0.4)

        jog = 0.0005
        for target in (cur + jog, cur - jog, cur):
            self._hold_position_command(part=part, target=target, hold_s=0.2)

        self.g.loggerUI.info(f"{part} recover mobility done @ {cur:.6f}")
        print("Recovery done")
        input("Press Enter to return")

    @hide_ui_while
    def auto_open_to_max(self, part: str, pos_name: str = "gripper_pos"):
        """
        自动持续张开夹爪到最大位置 - 方案选择菜单
        """
        print("=== 自动张开夹爪到最大 - 选择方案 ===")
        print("1. 方案1: 基于位置变化检测")
        print("2. 方案2: 基于电机堵转/力矩异常检测")
        
        choice = input("选择方案 (1/2): ").strip()
        
        if choice == "1":
            return self._auto_open_by_position_change(part, pos_name)
        elif choice == "2":
            return self._auto_open_by_stall_detection(part, pos_name)
        else:
            print("无效选择")
            input("按回车返回")
            return None

    @hide_ui_while
    def _auto_open_by_position_change(self, part: str, pos_name: str = "gripper_pos"):
        """
        方案1: 基于位置变化检测自动张开
        当连续多次位置变化小于阈值时停止
        """
        print("=== 方案1: 基于位置变化检测自动张开 ===")
        
        step = self.manual_control_step
        check_interval = 0.1
        stable_threshold = 0.0001
        stable_count_required = 5
        max_duration = 120.0  # 延长到120秒
        
        print(f"步长: {step:.6f}, 检测间隔: {check_interval}s")
        print(f"停止条件: 连续{stable_count_required}次位置变化<{stable_threshold}")
        print(f"最大运行时间: {max_duration}秒")
        input("按回车开始...")
        
        start_time = time.time()
        stable_count = 0
        positions_history = []
        last_print_time = 0
        
        print("开始自动张开...")
        while time.time() - start_time < max_duration:
            cur = self._get_feedback_scalar(pos_name)
            if cur is None:
                print("无法获取当前位置")
                break
            
            new_pos = cur + step
            self.g.robot.set_actions({part: {"type": "position", "position": [new_pos]}})
            
            time.sleep(check_interval)
            actual = self._get_feedback_scalar(pos_name)
            
            if actual is not None:
                positions_history.append(actual)
                if len(positions_history) > stable_count_required:
                    positions_history.pop(0)
                
                if len(positions_history) >= stable_count_required:
                    max_diff = max(positions_history) - min(positions_history)
                    if max_diff < stable_threshold:
                        print(f"位置已稳定，变化量: {max_diff:.6f}")
                        stable_count += 1
                        if stable_count >= 3:
                            print("检测到夹爪已到达最大位置")
                            break
                    else:
                        stable_count = 0
                
                # 每秒打印一次位置
                current_time = time.time()
                if current_time - last_print_time >= 1.0:
                    elapsed = current_time - start_time
                    if positions_history:
                        variation = max(positions_history) - min(positions_history)
                    else:
                        variation = 0.0
                    print(f"[{elapsed:.1f}s] 当前位置: {actual:.6f}, 变化量: {variation:.6f}")
                    last_print_time = current_time
        
        elapsed = time.time() - start_time
        if elapsed >= max_duration:
            print(f"达到最大运行时间 {max_duration}秒，停止运动")
        
        return self._record_and_return_open_data(pos_name)

    @hide_ui_while
    def _auto_open_by_stall_detection(self, part: str, pos_name: str = "gripper_pos"):
        """
        方案2: 基于电机堵转/力矩异常检测自动张开
        当检测到电机无法继续运动时停止
        """
        print("=== 方案2: 基于电机堵转检测自动张开 ===")
        print("注意: 此方案需要电机反馈力矩数据支持")
        
        step = self.manual_control_step
        check_interval = 0.1
        max_duration = 120.0  # 延长到120秒
        stall_threshold = 0.00005  # 位置变化小于此值认为堵转
        stall_count_required = 10  # 连续多次检测才确认
        
        print(f"步长: {step:.6f}, 堵转阈值: {stall_threshold}")
        print(f"最大运行时间: {max_duration}秒")
        input("按回车开始...")
        
        start_time = time.time()
        last_pos = None
        stall_count = 0
        last_print_time = 0
        
        print("开始自动张开，检测堵转...")
        while time.time() - start_time < max_duration:
            cur = self._get_feedback_scalar(pos_name)
            if cur is None:
                break
            
            new_pos = cur + step
            self.g.robot.set_actions({part: {"type": "position", "position": [new_pos]}})
            
            time.sleep(check_interval)
            actual = self._get_feedback_scalar(pos_name)
            
            if actual is not None and last_pos is not None:
                pos_change = abs(actual - last_pos)
                if pos_change < stall_threshold:
                    stall_count += 1
                    if stall_count >= stall_count_required:
                        print(f"检测到电机堵转(连续{stall_count}次)，停止运动")
                        break
                else:
                    if stall_count > 0:
                        print(f"堵转检测重置，变化量: {pos_change:.6f}")
                    stall_count = 0
            
            last_pos = actual
            
            # 每秒打印一次位置
            current_time = time.time()
            if current_time - last_print_time >= 1.0:
                elapsed = current_time - start_time
                print(f"[{elapsed:.1f}s] 当前位置: {actual:.6f}, 堵转计数: {stall_count}")
                last_print_time = current_time
        
        elapsed = time.time() - start_time
        if elapsed >= max_duration:
            print(f"达到最大运行时间 {max_duration}秒，停止运动")
        
        return self._record_and_return_open_data(pos_name)

    def _record_and_return_open_data(self, pos_name: str):
        """记录并返回张开后的数据"""
        print("\n稳定2秒后读取数据...")
        time.sleep(2.0)
        
        final_pos = self._get_feedback_scalar(pos_name)
        laser_dist = self._get_feedback_scalar("real_distance")
        
        print(f"\n=== 自动张开完成 ===")
        print(f"rad1 = {final_pos:.6f}")
        print(f"laser_open = {laser_dist:.6f}")
        
        self.g.loggerUI.info(f"自动张开完成: rad1={final_pos:.6f}, laser_open={laser_dist:.6f}")
        
        # 保存到临时存储
        if not hasattr(self, '_calibration_temp'):
            self._calibration_temp = {}
        self._calibration_temp["rad1"] = final_pos
        self._calibration_temp["laser_open"] = laser_dist
        
        input("按回车返回菜单")
        return final_pos, laser_dist

    @hide_ui_while
    def auto_close_to_min(self, part: str, pos_name: str = "gripper_pos"):
        """
        自动持续闭合夹爪到最小位置
        """
        print("=== 自动闭合夹爪到最小 ===")
        
        step = self.manual_control_step
        check_interval = 0.1
        stable_threshold = 0.0001
        stable_count_required = 5
        max_duration = 30.0
        
        print(f"步长: {step:.6f}, 检测间隔: {check_interval}s")
        input("按回车开始自动闭合...")
        
        start_time = time.time()
        stable_count = 0
        positions_history = []
        
        print("开始自动闭合...")
        while time.time() - start_time < max_duration:
            cur = self._get_feedback_scalar(pos_name)
            if cur is None:
                break
            
            new_pos = cur - step  # 闭合是减小位置
            self.g.robot.set_actions({part: {"type": "position", "position": [new_pos]}})
            
            time.sleep(check_interval)
            actual = self._get_feedback_scalar(pos_name)
            
            if actual is not None:
                positions_history.append(actual)
                if len(positions_history) > stable_count_required:
                    positions_history.pop(0)
                
                if len(positions_history) >= stable_count_required:
                    max_diff = max(positions_history) - min(positions_history)
                    if max_diff < stable_threshold:
                        stable_count += 1
                        if stable_count >= 3:
                            print("检测到夹爪已闭合到位")
                            break
                    else:
                        stable_count = 0
        
        print("稳定2秒后读取数据...")
        time.sleep(2.0)
        
        final_pos = self._get_feedback_scalar(pos_name)
        laser_dist = self._get_feedback_scalar("real_distance")
        
        print(f"\n=== 自动闭合完成 ===")
        print(f"rad2 = {final_pos:.6f}")
        print(f"laser_close = {laser_dist:.6f}")
        
        self.g.loggerUI.info(f"自动闭合完成: rad2={final_pos:.6f}, laser_close={laser_dist:.6f}")
        
        # 更新临时存储
        if hasattr(self, '_calibration_temp'):
            self._calibration_temp["rad2"] = final_pos
            self._calibration_temp["laser_close"] = laser_dist
        else:
            self._calibration_temp = {
                "rad2": final_pos,
                "laser_close": laser_dist
            }
        
        input("按回车返回菜单")
        return final_pos, laser_dist

    @hide_ui_while
    def calculate_and_write_length_per_radian(self, part: str):
        """
        计算并写入 length_per_radian
        """
        print("=== 计算 length_per_radian ===")
        
        # 检查是否有临时数据
        if not hasattr(self, '_calibration_temp'):
            print("错误: 没有标定数据，请先执行张开和闭合步骤")
            input("按回车返回")
            return
        
        temp = self._calibration_temp
        rad1 = temp.get("rad1")
        rad2 = temp.get("rad2")
        laser_open = temp.get("laser_open")
        laser_close = temp.get("laser_close")
        
        if rad1 is None or rad2 is None:
            print("错误: 缺少 rad1 或 rad2 数据")
            input("按回车返回")
            return
        
        if laser_open is None or laser_close is None:
            print("错误: 缺少激光数据")
            input("按回车返回")
            return
        
        # 计算
        real_distance = laser_open - laser_close
        delta_rad = abs(rad1 - rad2)
        
        if delta_rad < 1e-9:
            print(f"错误: 角度变化过小 ({delta_rad:.10f})")
            input("按回车返回")
            return
        
        length_per_radian = real_distance / (2000.0 * delta_rad)
        length_per_radian_10 = round(length_per_radian, 10)
        
        print(f"\n计算结果:")
        print(f"  rad1 = {rad1:.6f}")
        print(f"  rad2 = {rad2:.6f}")
        print(f"  delta_rad = {delta_rad:.6f}")
        print(f"  laser_open = {laser_open:.6f}")
        print(f"  laser_close = {laser_close:.6f}")
        print(f"  real_distance = {real_distance:.6f}")
        print(f"  length_per_radian = {length_per_radian_10:.10f}")
        
        # 写入YAML
        try:
            yaml_path = self._write_gripper_yaml_params(
                part=part,
                length_per_radian=length_per_radian_10,
            )
            print(f"\n✓ 已写入 {yaml_path}")
            self.g.loggerUI.info(f"length_per_radian 计算完成: {length_per_radian_10:.10f}, 已写入 {yaml_path}")
            
            # 清理临时数据
            delattr(self, '_calibration_temp')
            
        except Exception as e:
            print(f"写入YAML失败: {e}")
            self.g.loggerUI.error(f"写入YAML失败: {e}")
        
        input("按回车返回菜单")

    @hide_ui_while
    def repeatability_position_accuracy_test(self, part: str, pos_name: str, wait_for_return: bool = True):
        repeats = 10
        targets = [0.005, 0.045]
        settle_s = 1.0
        timeout_s = 3.0

        print("开始重复定位精度测试")
        print(f"流程: {targets} 循环 {repeats} 次, 每次到位后停留 {settle_s}s 采样")
        rows = []

        for idx in range(repeats):
            print(f"\n第 {idx + 1}/{repeats} 轮")
            aborted = False
            for target in targets:
                ok = self._move_to_target(part=part, pos_name=pos_name, target=target, timeout_s=timeout_s)
                if not ok:
                    self.g.loggerUI.warn(f"移动到 {target} 超时, 本次测试提前结束")
                    print(f"移动到 {target} 超时, 本次测试提前结束")
                    aborted = True
                    break

                time.sleep(settle_s)
                with self.g.feedback_lock:
                    left_raw = getattr(self.g.feedbackData, "laser_left", None)
                    right_raw = getattr(self.g.feedbackData, "laser_right", None)
                    real_raw = getattr(self.g.feedbackData, "real_distance", None)

                left_val = self._coerce_scalar(left_raw)
                right_val = self._coerce_scalar(right_raw)
                real_val = self._coerce_scalar(real_raw)
                rows.append([target, left_val, right_val, real_val])
                print(f"目标 {target:.3f} -> A:{left_val} B:{right_val} real:{real_val}")
            if aborted:
                break

        if not rows:
            print("没有采集到有效数据")
            if wait_for_return:
                input("回车返回")
            return

        out_dir = self._get_output_dir()
        os.makedirs(out_dir, exist_ok=True)
        tstamp = time.strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(out_dir, f"repeatability_accuracy_{part}_{tstamp}.csv")

        with open(out_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "期望夹爪距离",
                "激光位移传感器A读数(mm)",
                "激光位移传感器B读数(mm)",
                "激光位移传感器行程(mm)",
            ])
            writer.writerows(rows)

        self.g.loggerUI.info("重复定位精度测试完成")
        print(f"\n测试完成, 已保存 CSV: {out_file}")
        if wait_for_return:
            input("回车返回")

    @hide_ui_while
    def absolute_position_accuracy_test(self, part: str, pos_name: str, wait_for_return: bool = True):
        targets_m = [0.005, 0.01, 0.015, 0.02, 0.025, 0.03, 0.035, 0.04, 0.045, 0.05]
        settle_s = 1.0
        timeout_s = 3.0

        print("开始绝对定位精度测试")
        print(f"目标位置: {targets_m}, 每个点到位后停留 {settle_s}s 采样")
        rows = []

        for target_m in targets_m:
            ok = self._move_to_target(part=part, pos_name=pos_name, target=target_m, timeout_s=timeout_s)
            if not ok:
                self.g.loggerUI.warn(f"移动到 {target_m} 超时, 本次测试提前结束")
                print(f"移动到 {target_m} 超时, 本次测试提前结束")
                break

            time.sleep(settle_s)
            with self.g.feedback_lock:
                left_raw = getattr(self.g.feedbackData, "laser_left", None)
                right_raw = getattr(self.g.feedbackData, "laser_right", None)
                real_raw = getattr(self.g.feedbackData, "real_distance", None)

            left_val = self._coerce_scalar(left_raw)
            right_val = self._coerce_scalar(right_raw)
            real_val = self._coerce_scalar(real_raw)
            expected_stroke_mm = target_m * 2000.0  # 期望行程=2*夹爪移动位置(m), 再换算为mm

            rows.append([target_m, expected_stroke_mm, left_val, right_val, real_val])
            print(
                f"目标 {target_m:.3f}m (期望行程 {expected_stroke_mm:.3f}mm) "
                f"-> A:{left_val} B:{right_val} real:{real_val}"
            )

        if not rows:
            print("没有采集到有效数据")
            if wait_for_return:
                input("回车返回")
            return

        out_dir = self._get_output_dir()
        os.makedirs(out_dir, exist_ok=True)
        tstamp = time.strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(out_dir, f"absolute_accuracy_{part}_{tstamp}.csv")

        with open(out_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "期望夹移动位置(m)",
                "期望行程(mm)",
                "激光位移传感器A读数(mm)",
                "激光位移传感器B读数(mm)",
                "激光位移传感器行程(mm)",
            ])
            writer.writerows(rows)

        self.g.loggerUI.info("绝对定位精度测试完成")
        print(f"\n测试完成, 已保存 CSV: {out_file}")
        if wait_for_return:
            input("回车返回")

    @hide_ui_while
    def one_dim_force_accuracy_test(self, part: str, wait_for_return: bool = True):
        gripper_part = getattr(self.g, "selected_gripper", None)
        if not gripper_part:
            print("未找到当前夹爪名称，无法执行一维力精度测试")
            if wait_for_return:
                input("回车返回")
            return

        start_pos = 0.0088
        end_pos = 0.01
        sample_count = 3
        old_step = self.manual_control_step
        self.manual_control_step = 0.00005
        rows = []

        print("=== 一维力精度测试 ===")
        print(f"1. 自动移动夹爪到 {start_pos:.6f}")
        ok = self._move_to_target(part=gripper_part, pos_name="gripper_pos", target=start_pos, timeout_s=5.0)
        if not ok:
            self.g.loggerUI.warn(f"一维力精度测试: 移动到 {start_pos:.6f} 超时，继续进入点按采样")
            print(f"移动到 {start_pos:.6f} 超时，继续进入点按采样")

        try:
            print()
            print("2. 点按模式")
            print("   W: 张开方向微调")
            print("   S: 闭合方向微调")
            print("   R: 输入当前数显推拉力计读数并记录夹爪一维力通道0/1")
            print("   Q: 退出测试")
            print(f"   点按步长: {self.manual_control_step:.5f}")
            print(f"   需要记录 {sample_count} 次。按 W/S/R/Q 操作，W/S 不会回显到终端。")
            print()
            while len(rows) < sample_count:
                key = self._read_hidden_terminal_key()
                if key is None:
                    continue

                if key in ("w", "s"):
                    current = self._get_feedback_scalar("gripper_pos")
                    if current is None:
                        print("当前夹爪位置无效，无法点按")
                        continue
                    direction = 1.0 if key == "w" else -1.0
                    new_pos = current + direction * self.manual_control_step
                    self.g.robot.set_actions({gripper_part: {"type": "position", "position": [new_pos]}})
                    print(f"\r当前位置 {current:.6f} -> 目标 {new_pos:.6f}    ", end="", flush=True)
                elif key == "r":
                    print()
                    gauge_value = self._prompt_one_dim_force_gauge_value(
                        sample_index=len(rows) + 1,
                        sample_count=sample_count,
                    )
                    if gauge_value is None:
                        continue
                    with self.g.feedback_lock:
                        force_tip = list(getattr(self.g.feedbackData, "force_tip", [None, None]))
                    ch0 = self._coerce_scalar(force_tip[0] if len(force_tip) > 0 else None)
                    ch1 = self._coerce_scalar(force_tip[1] if len(force_tip) > 1 else None)
                    rows.append([gauge_value, ch0, ch1])
                    print(f"已记录 {len(rows)}/{sample_count}: 推拉力计={gauge_value}, 通道0={ch0}, 通道1={ch1}")
                elif key == "q":
                    print("\n一维力精度测试已退出")
                    return

            out_dir = self._get_output_dir()
            os.makedirs(out_dir, exist_ok=True)
            tstamp = time.strftime("%Y%m%d_%H%M%S")
            out_file = os.path.join(out_dir, f"one_dim_force_accuracy_{part}_{tstamp}.csv")
            with open(out_file, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "测量用一维力传感器数据（N）",
                    "夹爪一维力通道0数据（N）",
                    "夹爪一维力通道1数据（N）",
                ])
                writer.writerows(rows)

            self.g.loggerUI.info("一维力精度测试完成")
            print(f"\n一维力精度测试完成, CSV: {out_file}")
            print(f"4. 自动移动夹爪到 {end_pos:.6f}，请先取出/移开数显推拉力计")
            ok = self._move_to_target(part=gripper_part, pos_name="gripper_pos", target=end_pos, timeout_s=5.0)
            if not ok:
                self.g.loggerUI.warn(f"一维力精度测试: 移动到 {end_pos:.6f} 超时")
                print(f"移动到 {end_pos:.6f} 超时，请人工确认夹爪位置")
            input("确认数显推拉力计已安全移开后，按回车继续")
        finally:
            self.manual_control_step = old_step
            if wait_for_return:
                input("一维力精度测试结束，回车返回")

    def _read_hidden_terminal_key(self, block: bool = False):
        try:
            import select
            import termios
            import tty
        except ImportError:
            raw = input("请输入操作键 W/S/R/Q: ").strip().lower()
            return raw[:1] if raw else None

        fd = sys.stdin.fileno()
        old_attrs = termios.tcgetattr(fd)
        new_attrs = termios.tcgetattr(fd)
        new_attrs[3] = new_attrs[3] & ~(termios.ECHO | termios.ICANON)
        new_attrs[6][termios.VMIN] = 0
        new_attrs[6][termios.VTIME] = 1
        try:
            termios.tcsetattr(fd, termios.TCSADRAIN, new_attrs)
            timeout = None if block else 0.1
            ready, _, _ = select.select([sys.stdin], [], [], timeout)
            if not ready:
                return None
            ch = sys.stdin.read(1)
            return ch.lower() if ch else None
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_attrs)

    def _prompt_one_dim_force_gauge_value(self, sample_index: int, sample_count: int):
        raw = input(f"请输入第 {sample_index}/{sample_count} 次数显推拉力计读数(N): ").strip()
        if not raw:
            return None
        try:
            return float(raw)
        except ValueError:
            print("输入无效，请按 R 重新记录并输入数字")
            return None

    def start_manual_control_1dof(self, data_name: str, part: str):
        self.manual_control_active = True
        self.manual_data_name = data_name
        self.manual_part = part
        self.g.loggerUI.info("进入1D手动控制模式,按W/S以控制夹爪移动,按Q退出")

    def stop_manual_control_1dof(self):
        self.manual_control_active = False
        self.g.loggerUI.info("退出1D手动控制模式")

    @hide_ui_while
    def set_manual_control_step(self):
        raw = input(f"输入点按步长(当前 {self.manual_control_step:.6f}): ").strip()
        if not raw:
            return
        step = float(raw)
        if step <= 0:
            print("步长必须大于0")
            input("回车返回")
            return
        self.manual_control_step = step
        self.g.loggerUI.info(f"点按步长已设置: {self.manual_control_step:.6f}")
        print(f"点按步长已设置: {self.manual_control_step:.6f}")
        input("回车返回")
    def manual_control_1dof_step(self):
        """UI每帧调用一次"""
        if not self.manual_control_active:
            return
        data_name = self.manual_data_name
        part = self.manual_part
        step = float(self.manual_control_step)
        key = None
        try:
            ch = self.g.ui.win_menu.getch()
            if ch != -1:
                key = chr(ch).lower()
        except:
            return

        if key is None:
            return

        value = getattr(self.g.feedbackData, data_name)
        current = value[0]

        if key == 'w':
            new_pos = current + step
            self.g.loggerUI.info(f"手动控制('q'退出): {data_name} 从 {current:.4f} -> {new_pos:.4f}")
        elif key == 's':
            new_pos = current - step
            self.g.loggerUI.info(f"手动控制('q'退出): {data_name} 从 {current:.4f} -> {new_pos:.4f}")
        elif key == 'q':
            self.stop_manual_control_1dof()
            return
        else:
            return

        self.g.robot.set_actions({part: {"type": "position", "position": [new_pos]}})


    def set_zero(self,part):
        self.g.robot.send_command(part,{"command":"set_zero"})
        time.sleep(1)
        self._recover_motion_after_zero(part=part, pos_name="gripper_pos")
        self.g.loggerUI.info(f"{part}已硬件设零，并完成可动性恢复")
    def set_zero_loadcell(self,part:str,ch=0):
        self.g.robot.send_command(part, {"command": "calibrate_zero","index":ch}) ## index 0 for left setzero
        time.sleep(1)
        self.g.loggerUI.info(f"{part}的通道{ch}已硬件设零")

    @hide_ui_while
    def set_zero_loadcell_all(self, part: str):
        self.set_zero_loadcell(part=part, ch=0)
        self.set_zero_loadcell(part=part, ch=1)
        print("一维力传感器通道0/1已整体置零")
        input("按回车返回")

    def _set_lasers_zero_core(self):
        """
        Laser distance sensors set-zero for left and right (no UI interaction).
        """
        # Pause laser reading to avoid read/write frame collisions.
        prev_pause = getattr(self.g, "_laser_pause", False)
        self.g._laser_pause = True
        time.sleep(0.1)
        try:
            self.g._ensure_lasers()
        except Exception as e:
            self.g.loggerUI.error(f"laser set_zero failed: {e}")
            self.g._laser_pause = prev_pause
            return None

        if not getattr(self.g, "laser_left", None) and not getattr(self.g, "laser_right", None):
            self.g.loggerUI.error("laser set_zero failed: no laser connected")
            self.g._laser_pause = prev_pause
            return None

        results = {}
        if getattr(self.g, "laser_left", None):
            try:
                self.g.loggerUI.info(
                    f"[LaserLeft] port={self.g.laser_left.port}, open={self.g.laser_left.serial is not None and self.g.laser_left.serial.is_open}"
                )
                results["left"] = self.g.laser_left.set_zero(debug=False)
            except Exception as e:
                results["left"] = f"error:{e}"
        if getattr(self.g, "laser_right", None):
            try:
                self.g.loggerUI.info(
                    f"[LaserRight] port={self.g.laser_right.port}, open={self.g.laser_right.serial is not None and self.g.laser_right.serial.is_open}"
                )
                results["right"] = self.g.laser_right.set_zero(debug=False)
            except Exception as e:
                results["right"] = f"error:{e}"

        self.g.loggerUI.info(f"laser set_zero results: {results}")
        time.sleep(0.2)
        self.g._laser_pause = prev_pause
        return results

    @hide_ui_while
    def set_lasers_zero(self):
        """
        Laser distance sensors set-zero for left and right.
        """
        self._set_lasers_zero_core()
        input("按回车返回")

    @hide_ui_while
    def calibrate_loadcell(self,part:str,ch:int):
        known_force = input("请输入已知的标定力值(单位N),回车继续:")
        self.g.robot.send_command(part,{"command":"calibrate_force","index":ch,"force":float(known_force)})
        self.g.loggerUI.info(f"{part}的通道{ch}已标定为{known_force}N")

    @hide_ui_while
    def calibrate_loadcell_all(self, part: str):
        known_force = float(input("请输入已知的标定力值(单位N),将同时标定通道0/1,回车继续:"))
        self.g.robot.send_command(part, {"command": "calibrate_force", "index": 0, "force": known_force})
        self.g.robot.send_command(part, {"command": "calibrate_force", "index": 1, "force": known_force})
        self.g.loggerUI.info(f"{part}的通道0/1已整体标定为{known_force}N")
        print(f"一维力传感器通道0/1已整体标定为 {known_force}N")
        input("按回车返回")

    @hide_ui_while
    def zero_and_calibrate_loadcell_flow(self, part: str):
        gripper_part = getattr(self.g, "selected_gripper", None)
        if not gripper_part:
            print("未找到当前夹爪名称，无法执行一维力标零与标定流程")
            input("按回车返回")
            return

        old_step = self.manual_control_step
        self.manual_control_step = 0.00005
        try:
            print("=== 一维力传感器标零与标定流程 ===")
            print("1. 自动移动夹爪到 0.008700")
            ok = self._move_to_target(part=gripper_part, pos_name="gripper_pos", target=0.0087, timeout_s=5.0)
            if not ok:
                self.g.loggerUI.warn("一维力标定流程: 移动到 0.008700 超时，继续标零")
                print("移动到 0.008700 超时，继续标零")

            print("2. 对一维力传感器通道0/1整体标零")
            self.set_zero_loadcell(part=part, ch=0)
            self.set_zero_loadcell(part=part, ch=1)
            print("通道0/1已整体标零")
            self._show_loadcell_channels_for(1.0, prefix="标零后当前一维力")

            print()
            print("3. 点按调整并输入标定力")
            print("   W: 张开方向微调")
            print("   S: 闭合方向微调")
            print("   R: 输入当前数显推拉力计读数，并同时标定通道0/1")
            print("   Q: 退出流程")
            print(f"   点按步长: {self.manual_control_step:.5f}")

            while True:
                key = self._read_hidden_terminal_key()
                if key is None:
                    print(f"\r当前一维力: {self._format_loadcell_channels()}    ", end="", flush=True)
                    continue
                if key in ("w", "s"):
                    current = self._get_feedback_scalar("gripper_pos")
                    if current is None:
                        print("当前夹爪位置无效，无法点按")
                        continue
                    direction = 1.0 if key == "w" else -1.0
                    new_pos = current + direction * self.manual_control_step
                    self.g.robot.set_actions({gripper_part: {"type": "position", "position": [new_pos]}})
                    print(f"\r当前位置 {current:.6f} -> 目标 {new_pos:.6f}    ", end="", flush=True)
                elif key == "r":
                    print()
                    known_force = input("请输入当前数显推拉力计读数(N),回车后同时标定通道0/1: ").strip()
                    if not known_force:
                        continue
                    force = float(known_force)
                    self.g.robot.send_command(part, {"command": "calibrate_force", "index": 0, "force": force})
                    self.g.robot.send_command(part, {"command": "calibrate_force", "index": 1, "force": force})
                    self.g.loggerUI.info(f"{part}的通道0/1已按流程整体标定为{force}N")
                    print(f"通道0/1已整体标定为 {force}N")
                    print("标定后等待传感器稳定，请观察当前通道值")
                    self._show_loadcell_channels_for(2.0, prefix="标定后当前一维力")
                    input("确认通道0/1读数稳定且接近数显推拉力计读数后，按回车继续")
                    break
                elif key == "q":
                    print("\n一维力标零与标定流程已退出")
                    return

            input("请先移开/卸载数显推拉力计，确认安全后按回车移动到 0.010000")
            print("4. 自动移动夹爪到 0.010000")
            ok = self._move_to_target(part=gripper_part, pos_name="gripper_pos", target=0.01, timeout_s=5.0)
            if not ok:
                self.g.loggerUI.warn("一维力标定流程: 移动到 0.010000 超时")
                print("移动到 0.010000 超时，请人工确认夹爪位置")
            input("标零与标定流程完成，按回车返回")
        finally:
            self.manual_control_step = old_step
    
    def get_limits(self,part):
        print("当前关节范围:",self.g.robot.send_command(part, {"command": "get_limit"}))
        input("回车以返回")

    @hide_ui_while
    def zero_limit_travel_test(self, part: str, pos_name: str = "gripper_pos", wait_for_return: bool = True):
        """
        零位行程和限位测试：
        - 记录零点/0.05位置激光测距（mm）
        - 记录零点误差、软限位行程误差（相对0.05）
        - 记录多个目标输入的实际gripper_pos（含reload后的值）
        """
        def read_pos():
            return self._get_feedback_scalar(pos_name)

        def read_laser():
            return self._get_feedback_scalar("real_distance")

        def move_and_wait(target: float, wait_s: float = 0.0):
            ok = self._move_to_target(part=part, pos_name=pos_name, target=target, timeout_s=5.0)
            if not ok:
                self.g.loggerUI.warn(f"move to {target:.6f} timeout")
            if wait_s > 0:
                time.sleep(wait_s)

        def reload_and_wait():
            if not hasattr(self.g, "reload_robot_from_yaml"):
                raise RuntimeError("reload_robot_from_yaml not available")
            self.g.reload_robot_from_yaml()
            time.sleep(0.5)

        try:
            # 0.0
            move_and_wait(0.0, wait_s=2.0)
            zero_laser_mm = read_laser()
            zero_pos = read_pos()

            # 0.05
            move_and_wait(0.05, wait_s=2.0)
            soft_laser_mm = read_laser()
            soft_pos = read_pos()

            # 0.055
            move_and_wait(0.055, wait_s=0.5)
            pos_0055 = read_pos()

            # reload after 0.055
            reload_and_wait()
            pos_0055_reload = read_pos()

            # -0.005
            move_and_wait(-0.005, wait_s=0.5)
            pos_neg_0005 = read_pos()

            # reload after -0.005
            reload_and_wait()
            pos_neg_0005_reload = read_pos()

            # errors (target value for soft limit is 0.05 per requirement)
            zero_error_mm = None if zero_laser_mm is None else (zero_laser_mm - 0.0)
            soft_travel_error_mm = None if soft_laser_mm is None else (soft_laser_mm - 0.05)
            length_per_radian = self._get_feedback_scalar("config_length_per_radian")

            print("\n虚位测量准备: 先移动到 0.020")
            move_and_wait(0.02, wait_s=1.0)

            print("\n=== 虚位测量 ===")
            print("请先往夹爪闭合方向按，稳定后按任意键采样 realdistance1")
            self._read_hidden_terminal_key(block=True)
            realdistance1 = read_laser()
            print(f"realdistance1 = {realdistance1}")
            print("请再往夹爪张开方向扩，稳定后按任意键采样 realdistance2")
            self._read_hidden_terminal_key(block=True)
            realdistance2 = read_laser()
            print(f"realdistance2 = {realdistance2}")
            backlash_mm = None
            if realdistance1 is not None and realdistance2 is not None:
                backlash_mm = realdistance2 - realdistance1

            # write CSV
            log_dir = Path(self._get_output_dir())
            log_dir.mkdir(exist_ok=True)
            tstamp = time.strftime("%Y%m%d_%H%M%S")
            csv_path = log_dir / f"zero_limit_test_{tstamp}_{part}.csv"
            headers = [
                "零点位置激光测距（mm）",
                "软限位置（0.05）激光测距（mm）",
                "零点误差（mm）",
                "软限位行程误差（mm）",
                "0.0输入的gripper_pos",
                "0.05的gripper_pos",
                "0.055的gripper_pos",
                "0.055断电重启的gripper_pos",
                "-0.005的gripper_pos",
                "-0.005断电重启的gripper_pos",
            ]
            row = [
                zero_laser_mm,
                soft_laser_mm,
                zero_error_mm,
                soft_travel_error_mm,
                zero_pos,
                soft_pos,
                pos_0055,
                pos_0055_reload,
                pos_neg_0005,
                pos_neg_0005_reload,
            ]
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerow(row)
                writer.writerow(["length per radian", "虚位（mm）"])
                writer.writerow([length_per_radian, backlash_mm])

            self.g.loggerUI.info("零位行程和限位测试完成")
            print(f"零位行程和限位测试已保存: {csv_path}")
            if wait_for_return:
                input("按回车返回")
        except Exception as e:
            self.g.loggerUI.error(f"零位行程和限位测试失败: {e}")
            print(f"零位行程和限位测试失败: {e}")
            if wait_for_return:
                input("按回车返回")
    
    def move_multiple_points(self):
        #从csv文件加载为dict
        filename = input("请输入点位文件名(例如points.csv):").strip()
        points = []
        try:
            with open(filename,'r') as f:
                reader = csv.reader(f)
                header = next(reader)
                for row_idx, row in enumerate(reader):
                    q = [float(v) for v in row]
                    if len(q) != 7:
                        input(f"第{row_idx}行不是7个关节角，实际是{len(q)}个")
                        return
                    points.append(np.array(q,dtype=float))
            print(f"成功加载点位，共{len(points)}个")
        except FileNotFoundError:
            print(f"文件未找到:{filename}")
            input("回车以返回")
            return

        current_q = np.array(self.g.feedbackData.QActual, dtype=float)

        for idx, target_q in enumerate(points):

            print(f"\n===== 点 {idx}: {current_q.tolist()} -> {target_q.tolist()} =====")

            # 构造两点路径
            path = np.vstack([current_q, target_q])

            # TOPP 插值
            sample_rate = 20
            step = 1 / sample_rate
            try:
                ts, qs, qds, qdds, duration = TOPP(
                    path,
                    [self.max_vel] * 7,
                    [self.max_acc] * 7,
                    step
                )
            except RuntimeError as e:
                input("轨迹规划失败:", e)
                return

            # 执行轨迹
            print("移动中...")
            start_time = time.time()
            for t, q in zip(ts, qs):
                self.g.robot.set_actions({
                    self.g.selected_arm: {"type": "position", "position": q.tolist()}
                })
                while time.time() - start_time < t:
                    time.sleep(0.001)

            # 停留 5 秒
            print(f"到达第 {idx} 个点，停留 3.5 秒")
            stay_start = time.time()
            while time.time() - stay_start < 3.0:
                t = time.time() - stay_start
                time.sleep(0.01)

            # 手动控制走下一个点
            # input(f"到达第 {idx} 个点，确认测量成功后回车前往下一个点位")
            # time.sleep(1)

            # 下一段起点 = 当前段目标点
            current_q = target_q.copy()

        print("\n所有点已执行完成。")
        input("回车以返回")
        return

    @hide_ui_while
    def step_move_repetitively(self,part:str,pos_name:str):
        '''
        step_move_repetitively 的 Docstring

        :part: self.g.selected_arm
        :type part: str
        :pos_name: getattr(self.g.feedbackData, "rb_time")
        :type pos_name: str
        '''
        # 执行轨迹前清空日志
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()
        # 减少反馈负担
        self.g.feedback_mode = self.g.feedback_mode.BASIC
        user_input_start = input("请输入循环起始目标位置 (逗号分隔) 或 'q' 退出: ").strip()
        if user_input_start.lower() == 'q':
            print("结束单步移动记录")
            return
        user_input_end = input("请输入循环终止目标位置 (逗号分隔) 或 'q' 退出: ").strip()
        if user_input_end.lower() == 'q':
            print("结束单步移动记录")
            return
        while True:
            user_input = input("请输入重复次数（≥1 的整数）：")
            try:
                user_input_num = int(user_input)
                if user_input_num >= 1:
                    break
                else:
                    print("❌ 请输入大于等于 1 的整数")
            except ValueError:
                print("❌ 输入无效，请输入整数")
        print(f"✅ 输入成功：{user_input_num}")
        raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
        timeout_s = float(raw_input) if raw_input else 3.0
        try:
            q_pos_start = [float(x.strip()) for x in user_input_start.split(',')]
            q_pos_end = [float(x.strip()) for x in user_input_end.split(',')]
            print(f"开始空载重复测试: {q_pos_start} <-> {q_pos_end} 循环")
            
            # 执行移动并记录
            self.g.record_flag.set()
            # 占位
            print("开始高频+低频记录...")
            print("Moving...\n")
            for repeat_idx in range(user_input_num):
                print(f"\n====== 第 {repeat_idx + 1} / {user_input_num} 次循环 ======")
                for q_pos in (q_pos_start,q_pos_end):
                    print(f"目标位置: {q_pos}")
                    start_time = time.monotonic()
                    while True:
                        # 使用锁保护读取操作
                        with self.g.feedback_lock:
                            cur_q = getattr(self.g.feedbackData, pos_name)
                            rb_time = getattr(self.g.feedbackData, "rb_time")

                        # 到达判定
                        if np.max(np.abs(cur_q - np.array(q_pos))) <= TOL:
                            print("到达！")
                            self.g.loggerUI.info("到达！")
                            break

                        # 超时判定
                        elif time.monotonic() - start_time > timeout_s:
                            print(f"走点在 {timeout_s}s 后超时退出")
                            self.g.loggerUI.warn(f"走点在 {timeout_s}s 后超时退出")
                            break

                        # 发送控制命令
                        self.g.robot.set_actions({
                            part: {
                                "type": "position",
                                "position": q_pos
                            }
                        })

                        # 记录低频日志
                        self.g.logger._record_lowfreq(
                            t=time.perf_counter(),
                            q=q_pos,
                            rb_time=rb_time,
                            feedback_pos=cur_q
                        )

                        time.sleep(1 / CONTROL_HZ)
            # 延迟结束
            extra_record = time.perf_counter()
            while time.perf_counter() - extra_record <= EXTRA_TIME:
                with self.g.feedback_lock:
                    cur_q = getattr(self.g.feedbackData, pos_name)
                    rb_time = getattr(self.g.feedbackData, "rb_time")
                self.g.logger._record_lowfreq(t=time.time(),q=q_pos,
                                        rb_time=rb_time,
                                        feedback_pos=cur_q)
                time.sleep(1 / CONTROL_HZ)
            self.g.record_flag.clear()
            print("高频记录结束，共收集高频数据", len(self.g.highfreq_log))
            print("低频记录结束，共收集低频数据", len(self.g.lowfreq_log))
            # 减少反馈负担
            self.g.feedback_mode = self.g.feedback_mode.FULL
            print("\n=== 数据记录 ===")
            confirm = input("是否保存日志? (y/n): ").strip().lower()
            if confirm != "y":
                print("日志保存更新取消")
                input("回车以返回")
                return

            # 保存日志
            highfreq_filename, lowfreq_filename,tstamp = self.g.logger._save_logs(part)

            confirm = input("是否可视化记录数据? (y/n): ").strip().lower()
            if confirm != "y":
                print("数据可视化取消")
                input("回车以返回")
                return

            draw(log_dir=f'{project_root}/logs',lowfile=lowfreq_filename,highfile=highfreq_filename,savefig=f"viz_{tstamp}.png")

        except ValueError:
            print("输入无效，请输入逗号分隔的数字列表或 'exit' 退出。")

    @hide_ui_while
    def grasp_test(self, part: str, pos_name: str):
        from collections import deque
        import matplotlib.pyplot as plt
        import time
        import numpy as np
        print("[堵转测试步骤1] 移动夹爪到未抓取/堵转状态位置")

        raw_input = input(f"以逗号分隔的方式输入N自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N) 或 'q' 退出: ")
        if raw_input.lower() == 'q':
            return False
        q_pos = [float(x.strip()) for x in raw_input.split(',')]
        dof = 1
        if len(q_pos) != dof:
            print(f"\nError: Expected {dof} values, but got {len(q_pos)}.")
            return False
        raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
        timeout_s = float(raw_input) if raw_input else 3.0
        print(f"\n移动到新位置: {[f'{p:.3f}' for p in q_pos]}")
        self._run_point(q_name=pos_name,q_pos=q_pos,part=part,timeout_s=timeout_s)

        time.sleep(0.1)
        with self.g.feedback_lock:
            record_tip_1 = np.array(getattr(self.g.feedbackData, "force_tip"))
            record_temperature_1 = np.array(getattr(self.g.feedbackData, "temperature"))

        user_input = input("[堵转测试步骤2] 请输入目标位置 (逗号分隔) 或 'q' 退出: ").strip()
        if user_input.lower() == 'q':
            return

        raw_input = input("设置超时时间(秒),默认3秒，回车继续: ")
        timeout_s = float(raw_input) if raw_input else 3.0

        fig = None
        try:
            # 根据超时时间配置绘图缓存长度，避免只显示固定时长
            max_points = int(max(1000, (timeout_s + 5.0) * CONTROL_HZ))

            buf_time = deque(maxlen=max_points)
            buf_q = deque(maxlen=max_points)
            buf_temp = deque(maxlen=max_points)
            buf_torque = deque(maxlen=max_points)
            buf_force_0 = deque(maxlen=max_points)
            buf_force_1 = deque(maxlen=max_points)

            # ===== 主线程创建图（关键）=====
            plt.ion()
            fig, axes = plt.subplots(4, 1, figsize=(8, 8), sharex=True)

            ax_q, ax_temp, ax_torque, ax_force = axes

            line_q, = ax_q.plot([], [], 'b-', label="Position")
            line_temp, = ax_temp.plot([], [], 'r-', label="Temperature")
            line_torque, = ax_torque.plot([], [], 'g-', label="Torque")
            line_force_0, = ax_force.plot([],[],'c-',label="Tip[0]")
            line_force_1, = ax_force.plot([],[],'m-',label="Tip[1]")
            ax_q.set_ylabel("Position")
            ax_temp.set_ylabel("Temperature")
            ax_torque.set_ylabel("Torque")
            ax_torque.set_xlabel("Time (s)")
            ax_force.set_ylabel("Tip Force")
            # ===== 数值显示（右上角）=====
            txt_q = ax_q.text(
                0.98, 0.95, "",
                transform=ax_q.transAxes,
                ha="right", va="top"
            )

            txt_temp = ax_temp.text(
                0.98, 0.95, "",
                transform=ax_temp.transAxes,
                ha="right", va="top"
            )

            txt_torque = ax_torque.text(
                0.98, 0.95, "",
                transform=ax_torque.transAxes,
                ha="right", va="top"
            )

            txt_force = ax_force.text(
                0.98, 0.95, "",
                transform=ax_force.transAxes,
                ha="right", va="top"
            )
            for ax in axes:
                ax.grid(True)
                ax.legend()

            last_plot_time = 0.0
            PLOT_DT = 0.05  # 20 Hz 刷新

            # ===== 清空日志 =====
            self.g.highfreq_log.clear()
            self.g.lowfreq_log.clear()

            q_pos = np.array([float(x.strip()) for x in user_input.split(',')])
            print(f"移动到新位置: {[f'{p:.3f}' for p in q_pos]}")

            self.g.record_flag.set()
            print("开始高频+低频记录...\n")

            start_time = time.monotonic()

            while True:
                with self.g.feedback_lock:
                    cur_q = np.array(getattr(self.g.feedbackData, pos_name))
                    cur_temperature = float(getattr(self.g.feedbackData, "temperature"))
                    cur_torque = float(getattr(self.g.feedbackData, "gripper_torque"))
                    cur_tip_list = np.array(getattr(self.g.feedbackData, "force_tip"))

                # ===== 控制逻辑 =====
                if np.max(np.abs(cur_q - q_pos)) <= TOL:
                    print("到达！")
                    break
                elif time.monotonic() - start_time > timeout_s:
                    print(f"走点在 {timeout_s}s 后超时退出")
                    break

                self.g.robot.set_actions({part: {"type": "position", "position": q_pos.tolist()}})

                # ===== 记录数据 =====
                now = time.monotonic()
                buf_time.append(now)
                buf_q.append(cur_q[0])           # 单自由度
                buf_temp.append(cur_temperature)
                buf_torque.append(cur_torque)
                buf_force_0.append(cur_tip_list[0])
                buf_force_1.append(cur_tip_list[1])

                # ===== 刷新图像（主线程，安全）=====
                if now - last_plot_time > PLOT_DT and len(buf_time) > 2:
                    last_plot_time = now
                    t0 = buf_time[0]
                    t = [x - t0 for x in buf_time]

                    line_q.set_data(t, list(buf_q))
                    line_temp.set_data(t, list(buf_temp))
                    line_torque.set_data(t, list(buf_torque))
                    line_force_0.set_data(t, list(buf_force_0))
                    line_force_1.set_data(t, list(buf_force_1))
                    txt_q.set_text(f"{buf_q[-1]:.4f}")
                    txt_temp.set_text(f"{buf_temp[-1]:.2f} °C")
                    txt_torque.set_text(f"{buf_torque[-1]:.3f}")

                    txt_force.set_text(
                        f"Tip0: {buf_force_0[-1]:.3f}\n"
                        f"Tip1: {buf_force_1[-1]:.3f}"
                    )

                    for ax in axes:
                        ax.relim()
                        ax.autoscale_view()

                    plt.pause(0.001)  # 驱动 GUI

                time.sleep(1 / CONTROL_HZ)

            self.g.record_flag.clear()
            print("高频记录结束，共收集高频数据", len(self.g.highfreq_log))

            confirm = input("是否保存日志? (y/n): ").strip().lower()
            if confirm == "y":
                self.g.logger._save_logs(part)

            print("[堵转测试步骤3] 移动夹爪到未抓取/堵转状态位置")
            self._run_point_without_hide(q_name="gripper_pos",dof=1,part=part)
            time.sleep(0.1)
            with self.g.feedback_lock:
                record_tip_2 = np.array(getattr(self.g.feedbackData, "force_tip"))
                record_temperature_2 = np.array(getattr(self.g.feedbackData, "temperature"))

            self.g.loggerUI.info(f"[测试结果] 堵转前一维力:{record_tip_1}，堵转后一维力:{record_tip_2}，堵转时间:{timeout_s}")
            self.g.loggerUI.info(f"[测试结果] 堵转前电机温度:{record_temperature_1}，堵转后电机温度:{record_temperature_2},堵转时间:{timeout_s}")

        except Exception as e:
            print(e)

        finally:
            if fig is not None:
                try:
                    tstamp = time.strftime("%Y%m%d_%H%M%S")
                    save_path = f"{project_root}/logs/grasp_test_{tstamp}_{part}.png"
                    fig.savefig(save_path, dpi=200, bbox_inches='tight')
                    print(f"图已保存: {save_path}")
                except Exception as e:
                    print(f"图保存失败: {e}")
            plt.ioff()
            if fig is not None:
                plt.close(fig)

    @hide_ui_while
    def integrity_test(self, part: str, pos_name: str):
        from collections import deque
        import matplotlib.pyplot as plt
        import time
        import numpy as np
        print("[测试步骤1] 移动夹爪到未抓取/堵转状态位置")
        
        raw_input = input(f"以逗号分隔的方式输入N自由度的期望移动位置 (e.g. 0.1 or 0.1,0.2,...,N) 或 'q' 退出: ")
        if raw_input.lower() == 'q':
            return False
        q_pos = [float(x.strip()) for x in raw_input.split(',')]
        dof = 1
        if len(q_pos) != dof:
            print(f"\nError: Expected {dof} values, but got {len(q_pos)}.")
            return False
        raw_input = input(f"设置超时时间(秒),默认3秒，回车继续: ")
        timeout_s = float(raw_input) if raw_input else 3.0
        print(f"\n移动到新位置: {[f'{p:.3f}' for p in q_pos]}")
        self._run_point(q_name=pos_name,q_pos=q_pos,part=part,timeout_s=timeout_s)

        time.sleep(0.1)
        with self.g.feedback_lock:
            record_tip_1 = np.array(getattr(self.g.feedbackData, "force_tip"))
            record_temperature_1 = np.array(getattr(self.g.feedbackData, "temperature"))

        MAX_POINTS = 1500

        buf_time = deque(maxlen=MAX_POINTS)
        buf_q = deque(maxlen=MAX_POINTS)
        buf_temp = deque(maxlen=MAX_POINTS)
        buf_torque = deque(maxlen=MAX_POINTS)
        buf_force_0 = deque(maxlen=MAX_POINTS)
        buf_force_1 = deque(maxlen=MAX_POINTS)

        # ===== 主线程创建图（关键）=====
        plt.ion()
        fig, axes = plt.subplots(4, 1, figsize=(8, 8), sharex=True)

        ax_q, ax_temp, ax_torque, ax_force = axes

        line_q, = ax_q.plot([], [], 'b-', label="Position")
        line_temp, = ax_temp.plot([], [], 'r-', label="Temperature")
        line_torque, = ax_torque.plot([], [], 'g-', label="Torque")
        line_force_0, = ax_force.plot([],[],'c-',label="Tip[0]")
        line_force_1, = ax_force.plot([],[],'m-',label="Tip[1]")
        ax_q.set_ylabel("Position")
        ax_temp.set_ylabel("Temperature")
        ax_torque.set_ylabel("Torque")
        ax_torque.set_xlabel("Time (s)")
        ax_force.set_ylabel("Tip Force")
        # ===== 数值显示（右上角）=====
        txt_q = ax_q.text(
            0.98, 0.95, "",
            transform=ax_q.transAxes,
            ha="right", va="top"
        )

        txt_temp = ax_temp.text(
            0.98, 0.95, "",
            transform=ax_temp.transAxes,
            ha="right", va="top"
        )

        txt_torque = ax_torque.text(
            0.98, 0.95, "",
            transform=ax_torque.transAxes,
            ha="right", va="top"
        )

        txt_force = ax_force.text(
            0.98, 0.95, "",
            transform=ax_force.transAxes,
            ha="right", va="top"
        )
        for ax in axes:
            ax.grid(True)
            ax.legend()

        last_plot_time = 0.0
        PLOT_DT = 0.05  # 20 Hz 刷新

        # ===== 清空日志 =====
        self.g.highfreq_log.clear()
        self.g.lowfreq_log.clear()

        # 循环输入起始位置
        while True:
            user_input_start = input("[测试步骤2] 请输入循环起始位置 (逗号分隔) 或 'q' 退出: ").strip()
            if user_input_start.lower() == 'q':
                plt.close(fig)
                return
            try:
                start_pos = np.array([float(x.strip()) for x in user_input_start.split(',')])
                break  # 输入合法，跳出循环
            except ValueError:
                print("⚠️ 输入格式错误，请用逗号分隔数字，如 0.1,0.2,...")

        # 循环输入终止位置
        while True:
            user_input_end = input("请输入循环终止位置 (逗号分隔) 或 'q' 退出: ").strip()
            if user_input_end.lower() == 'q':
                plt.close(fig)
                return
            try:
                end_pos = np.array([float(x.strip()) for x in user_input_end.split(',')])
                break  # 输入合法，跳出循环
            except ValueError:
                print("⚠️ 输入格式错误，请用逗号分隔数字，如 0.1,0.2,...")


        raw_input = input("设置超时时间（秒）默认3秒，回车继续：")
        timeout_s = float(raw_input) if raw_input else 3.0
        while True:
            user_input = input("请输入重复次数（≥1 的整数）：")
            try:
                user_input_num = int(user_input)
                if user_input_num >= 1:
                    break
                else:
                    print("❌ 请输入大于等于 1 的整数")
            except ValueError:
                print("❌ 输入无效，请输入整数")
        print(f"✅ 输入成功：{user_input_num}")
        try:
            q_pos_start = np.array([float(x) for x in user_input_start.split(',')])
            q_pos_end   = np.array([float(x) for x in user_input_end.split(',')])

            self.g.record_flag.set()
            print("开始高频+低频记录...\n")

            for repeat_idx in range(user_input_num):
                print(f"\n====== 第 {repeat_idx+1}/{user_input_num} 次堵转循环 ======")
                
                for q_pos in (q_pos_start, q_pos_end):
                    print(f"目标位置: {q_pos.tolist()}")
                    start_time = time.monotonic()

                    while True:
                        with self.g.feedback_lock:
                            cur_q = np.array(getattr(self.g.feedbackData, pos_name))
                            cur_temperature = float(getattr(self.g.feedbackData, "temperature"))
                            cur_torque = float(getattr(self.g.feedbackData, "gripper_torque"))
                            cur_tip_list = np.array(getattr(self.g.feedbackData, "force_tip"))

                        # ===== 控制逻辑 =====
                        if np.max(np.abs(cur_q - q_pos)) <= TOL:
                            print("到达！")
                            break
                        elif time.monotonic() - start_time > timeout_s:
                            print(f"走点在 {timeout_s}s 后超时退出")
                            break

                        self.g.robot.set_actions({part: {"type": "position", "position": q_pos.tolist()}})

                        # ===== 记录数据 =====
                        now = time.monotonic()
                        buf_time.append(now)
                        buf_q.append(cur_q[0])           # 单自由度
                        buf_temp.append(cur_temperature)
                        buf_torque.append(cur_torque)
                        buf_force_0.append(cur_tip_list[0])
                        buf_force_1.append(cur_tip_list[1])

                        # ===== 刷新图像（主线程，安全）=====
                        if now - last_plot_time > PLOT_DT and len(buf_time) > 2:
                            last_plot_time = now
                            t0 = buf_time[0]
                            t = [x - t0 for x in buf_time]

                            line_q.set_data(t, list(buf_q))
                            line_temp.set_data(t, list(buf_temp))
                            line_torque.set_data(t, list(buf_torque))
                            line_force_0.set_data(t, list(buf_force_0))
                            line_force_1.set_data(t, list(buf_force_1))
                            txt_q.set_text(f"{buf_q[-1]:.4f}")
                            txt_temp.set_text(f"{buf_temp[-1]:.2f} °C")
                            txt_torque.set_text(f"{buf_torque[-1]:.3f}")

                            txt_force.set_text(
                                f"Tip0: {buf_force_0[-1]:.3f}\n"
                                f"Tip1: {buf_force_1[-1]:.3f}"
                            )

                            for ax in axes:
                                ax.relim()
                                ax.autoscale_view()

                            plt.pause(0.001)  # 驱动 GUI

                        time.sleep(1 / CONTROL_HZ)

            self.g.record_flag.clear()
            print("高频记录结束，共收集高频数据", len(self.g.highfreq_log))

            confirm = input("是否保存日志? (y/n): ").strip().lower()
            if confirm == "y":
                self.g.logger._save_logs(part)

            print("[测试步骤3] 移动夹爪到未抓取/堵转状态位置")
            self._run_point_without_hide(q_name="gripper_pos",dof=1,part=part)
            time.sleep(0.1)
            with self.g.feedback_lock:
                record_tip_2 = np.array(getattr(self.g.feedbackData, "force_tip"))
                record_temperature_2 = np.array(getattr(self.g.feedbackData, "temperature"))

            self.g.loggerUI.info(f"[测试结果] 堵转前一维力:{record_tip_1}，堵转后一维力:{record_tip_2}，堵转时间:{timeout_s}")
            self.g.loggerUI.info(f"[测试结果] 堵转前电机温度:{record_temperature_1}，堵转后电机温度:{record_temperature_2},堵转时间:{timeout_s}")

        except Exception as e:
            print(e)

        finally:
            plt.ioff()
            plt.close(fig)
















